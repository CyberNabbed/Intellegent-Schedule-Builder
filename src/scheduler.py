import customtkinter as ctk
from tkinter import messagebox, filedialog
import sys
import os
import calendar
from datetime import datetime
import re
from openpyxl.styles import PatternFill, Font, Alignment
from ortools.sat.python import cp_model

# Ensure imports work
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from schedule_engine_v2 import ScheduleGeneratorV2, ALL_SHIFTS, SHIFT_NAMES, PHONE_SHIFTS, FD_SHIFTS, P1, P2, P3, P4, FD1, FD2
from config_manager import ConfigManager
from ics_export import (
    generate_all_ics, is_outlook_available, send_via_outlook,
    build_email_body, build_agent_events
)

# ── Appearance ────────────────────────────────────────────────────────────────
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("dark-blue")

# ── Design Tokens ─────────────────────────────────────────────────────────────
# Accent / brand colour (blue that works in both modes)
ACCENT        = "#3B82F6"          # blue-500
ACCENT_HOVER  = "#2563EB"          # blue-600
ACCENT_LIGHT  = ("#DBEAFE", "#1E3A5F")  # blue-100 / dark tint

# Sidebar
SIDEBAR_BG    = ("#F8FAFC", "#1A1F2E")   # near-white / dark-navy

# Surface colours for cards / panel backgrounds
SURFACE       = ("#FFFFFF", "#242938")   # white / slightly lighter navy
SURFACE_ALT   = ("#F1F5F9", "#1E2333")   # very light grey / mid-navy

# Text
TEXT_PRIMARY  = ("#0F172A", "#F1F5F9")   # near-black / near-white
TEXT_MUTED    = ("#64748B", "#94A3B8")   # slate-500 / slate-400
TEXT_ON_ACCENT = "white"

# Status colours
RED_BG    = ("#FEE2E2", "#7F1D1D")   # red-100 / red-900
RED_TEXT  = ("#DC2626", "#FCA5A5")   # red-600 / red-300
ORANGE_BG = ("#FEF3C7", "#451A03")   # amber-100 / amber-900
ORANGE_TEXT = ("#D97706", "#FCD34D") # amber-600 / amber-300
BLUE_BG   = ("#DBEAFE", "#1E3A5F")   # blue-100 / dark tint
BLUE_TEXT = ("#1D4ED8", "#93C5FD")   # blue-700 / blue-300
GREEN_BG  = ("#D1FAE5", "#064E3B")   # green-100 / green-900

# Cell colours (schedule table)
CELL_PHONE   = ("#DBEAFE", "#1E3A5F")  # blue tint
CELL_FD      = ("#FEF3C7", "#3D2906")  # amber tint
CELL_ERROR   = ("#FEE2E2", "#7F1D1D")  # red tint
CELL_EMPTY   = ("#F8FAFC", "#1A1F2E")  # neutral background
CELL_HEADER  = ACCENT

# Border colour
BORDER       = ("#E2E8F0", "#2D3548")

# Scrollbar colour (subtle by default, becomes ACCENT on hover)
SCROLLBAR    = ("#F1F5F9", "#1E2333")  # Very subtle, matches surface

# Radius
RADIUS_SM    = 6
RADIUS_MD    = 10
RADIUS_LG    = 14

# Font sizes
FS_XS   = 13
FS_SM   = 14
FS_BASE = 15
FS_LG   = 17
FS_XL   = 20
FS_2XL  = 24
FS_3XL  = 30


def make_font(size=FS_BASE, weight="normal"):
    return ctk.CTkFont(size=size, weight=weight)


def section_label(parent, text, **kw):
    """Uppercase muted section heading used throughout the sidebar."""
    return ctk.CTkLabel(
        parent, text=text.upper(),
        font=make_font(FS_XS, "bold"),
        text_color=TEXT_MUTED,
        anchor="w", **kw
    )


def divider_pack(parent, pady=(8, 12)):
    """Thin separator for pack-managed parents."""
    f = ctk.CTkFrame(parent, height=1, fg_color=BORDER, corner_radius=0)
    f.pack(fill="x", pady=pady)
    return f


def divider_grid(parent, row, pady=(8, 12)):
    """Thin separator for grid-managed parents."""
    f = ctk.CTkFrame(parent, height=1, fg_color=BORDER, corner_radius=0)
    f.grid(row=row, column=0, sticky="ew", pady=pady)
    return f


# ── Parsing Helpers (pure functions, testable without GUI) ────────────────────

def parse_holidays_string(raw_text, max_day, log_fn=None):
    """Parse a comma-separated holiday string into a list of day numbers.

    Supports single days and ranges (e.g. "1, 5, 10-15").
    Returns a list of ints.
    """
    holidays = []
    if not raw_text:
        return holidays
    for p in raw_text.split(','):
        p = p.strip()
        if not p:
            continue
        if '-' in p:
            try:
                start, end = map(int, p.split('-'))
                if start < 1 or end > max_day or start > end:
                    if log_fn:
                        log_fn(f"  ⚠  Invalid holiday range: {p}")
                    continue
                holidays.extend(range(start, end + 1))
            except ValueError:
                if log_fn:
                    log_fn(f"  ⚠  Invalid holiday: {p}")
        else:
            try:
                day = int(p)
                if day < 1 or day > max_day:
                    if log_fn:
                        log_fn(f"  ⚠  Holiday out of range: {p}")
                    continue
                holidays.append(day)
            except ValueError:
                if log_fn:
                    log_fn(f"  ⚠  Invalid holiday: {p}")
    return holidays


def parse_timeoff_entries(entries_dict, dates_list, staff_names, max_day, log_fn=None):
    """Parse time-off entry strings into a dict of {(name, day_index): True}.

    entries_dict: {name: raw_text_string}
    dates_list: list of date objects from the generator
    staff_names: list of valid employee names
    max_day: max day of month
    Returns dict suitable for passing to build_model().
    """
    to_requests = {}
    for name, raw_text in entries_dict.items():
        if not raw_text:
            continue
        for d_str in re.split(r'[ ,/]+', raw_text):
            if not d_str:
                continue
            if '-' in d_str:
                try:
                    start, end = map(int, d_str.split('-'))
                    if start < 1 or end > max_day or start > end:
                        if log_fn:
                            log_fn(f"  ⚠  Invalid range for {name}: {d_str}")
                        continue
                    days_added = []
                    days_skipped = []
                    for day_num in range(start, end + 1):
                        found = False
                        for idx, dt in enumerate(dates_list):
                            if dt.day == day_num:
                                if name in staff_names:
                                    to_requests[(name, idx)] = True
                                    days_added.append(day_num)
                                found = True
                                break
                        if not found:
                            days_skipped.append(day_num)
                    if log_fn:
                        if days_added:
                            log_fn(f"  ✓  {name}: off days {start}-{end}")
                        if days_skipped:
                            log_fn(f"  ⚠  {name}: days {','.join(map(str, days_skipped))} skipped (weekend/holiday)")
                except ValueError:
                    if log_fn:
                        log_fn(f"  ⚠  Invalid range for {name}: {d_str}")
            else:
                try:
                    day_num = int(d_str)
                    if day_num < 1 or day_num > max_day:
                        if log_fn:
                            log_fn(f"  ⚠  Bad day for {name}: {day_num}")
                        continue
                    found = False
                    for idx, dt in enumerate(dates_list):
                        if dt.day == day_num:
                            if name in staff_names:
                                to_requests[(name, idx)] = True
                                if log_fn:
                                    log_fn(f"  ✓  {name}: off {dt.strftime('%a %b %d')}")
                            found = True
                            break
                    if not found and log_fn:
                        log_fn(f"  ⚠  {name}: day {day_num} skipped (weekend/holiday)")
                except ValueError:
                    if log_fn:
                        log_fn(f"  ⚠  Invalid input for {name}: {d_str}")
    return to_requests


def parse_advanced_options(start_text, full_weeks, max_day, log_fn=None):
    """Parse advanced scheduling options.

    Returns (generate_full_weeks: bool, start_from_day: int|None).
    """
    start_from_day = None
    if start_text:
        try:
            start_from_day = int(start_text)
            if start_from_day < 1 or start_from_day > max_day:
                if log_fn:
                    log_fn(f"  ⚠  Invalid start day: {start_from_day}. Using default (1st).")
                start_from_day = None
            else:
                if log_fn:
                    log_fn(f"  Starting schedule from day {start_from_day}")
        except ValueError:
            if log_fn:
                log_fn(f"  ⚠  Invalid start day input: '{start_text}'. Using default (1st).")
            start_from_day = None

    if full_weeks and log_fn:
        log_fn("  Generating full calendar weeks (may extend into next month)")

    return full_weeks, start_from_day


def map_week_col_to_engine_day(generator, week_idx, col_idx):
    """Map preview grid coordinates to generator day index.
    Matches the exact week-switching logic in schedule_engine_v2.py:get_weekly_matrix().
    """
    if not generator:
        return None
    
    current_week_idx = -1
    last_weekday = -1
    
    for d_idx, dt in enumerate(generator.dates):
        wd = dt.weekday()  # 0=Mon, 4=Fri
        if wd > 4:
            continue
            
        # Match get_weekly_matrix logic for starting a new week
        if current_week_idx == -1 or wd == 0 or wd < last_weekday:
            current_week_idx += 1
            
        last_weekday = wd
        
        if current_week_idx == week_idx and wd == col_idx:
            return d_idx
            
    return None


class MockSolver:
    """Mock solver that returns saved schedule values instead of solving."""
    def __init__(self, schedule_data, shifts_dict):
        """
        schedule_data: dict mapping "day,shift" -> employee_name
        shifts_dict: the generator.shifts dict mapping (emp, day, shift) -> BoolVar
        """
        self.schedule_data = schedule_data
        self.shifts_dict = shifts_dict
        # Create reverse lookup: BoolVar -> (emp, day, shift)
        self.var_to_key = {var: key for key, var in shifts_dict.items()}

    def Value(self, shift_var):
        """Mock the Value() method to return saved assignments."""
        # Look up which (emp, day, shift) this variable represents
        key = self.var_to_key.get(shift_var)
        if key is None:
            return 0

        emp, day, shift = key
        schedule_key = f"{day},{shift}"
        assigned_emp = self.schedule_data.get(schedule_key, "")
        return 1 if assigned_emp == emp else 0


class ToolTip:
    """
    Lightweight hover tooltip for any widget.
    Shows a small popup with descriptive text after a brief delay.
    """
    DELAY_MS = 400

    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self._tip_window = None
        self._after_id = None
        widget.bind("<Enter>", self._schedule, add="+")
        widget.bind("<Leave>", self._hide, add="+")
        widget.bind("<ButtonPress>", self._hide, add="+")

    def _schedule(self, event=None):
        if not self.widget.winfo_exists():
            return
        self._after_id = self.widget.after(self.DELAY_MS, self._show)

    def _show(self):
        if self._tip_window or not self.text or not self.widget.winfo_exists():
            return
        # Position: below the widget, slightly offset
        x = self.widget.winfo_rootx() + 10
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 4
        import tkinter as tk
        tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        # Style the tooltip
        frame = tk.Frame(tw, background="#1E293B", bd=0, relief="flat")
        frame.pack()
        label = tk.Label(
            frame, text=self.text,
            background="#1E293B", foreground="#F1F5F9",
            font=("TkDefaultFont", 10),
            wraplength=280,
            justify="left",
            padx=10, pady=6
        )
        label.pack()
        self._tip_window = tw

    def _hide(self, event=None):
        if self._after_id and self.widget.winfo_exists():
            self.widget.after_cancel(self._after_id)
            self._after_id = None
        if self._tip_window and self._tip_window.winfo_exists():
            self._tip_window.destroy()
            self._tip_window = None


# ── Main Application ──────────────────────────────────────────────────────────
class ScheduleAppV2(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Schedule Builder")
        self.geometry("1180x820")
        self.minsize(900, 680)

        self.config_manager = ConfigManager()
        self.employees = self.config_manager.get_employees()
        self.generator = None
        self.edited_assignments = {}  # Track manual edits: (week_idx, day_idx, shift_idx) -> name
        self.cell_widgets = {}  # Store cell widget references: (week_idx, day_idx, shift_idx) -> label widget
        self.cell_violations = {}  # Track violations: (week_idx, day_idx, shift_idx) -> (violation_type, conflicting_cell_key)
        self.timeoff_entries = {}
        self.cached_timeoff_requests = {}  # Snapshot of time-off when schedule was generated: {(name, day_idx): True}

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        if not self.employees:
            self.show_wizard()
        else:
            self.show_dashboard()

    # ── Helpers ───────────────────────────────────────────────────────────────

    def clear_window(self):
        # Unregister appearance tracker so it doesn't fire against destroyed widgets
        try:
            ctk.AppearanceModeTracker.remove(self._on_appearance_change)
        except Exception:
            pass
        for widget in self.winfo_children():
            widget.destroy()

    def validate_employee_name(self, name, existing_names=None):
        if not name or not name.strip():
            return False, "Name cannot be empty."
        name = name.strip()
        if len(name) > 50:
            return False, "Name cannot exceed 50 characters."
        if not re.match(r"^[a-zA-Z][a-zA-Z\s'\-]{0,49}$", name):
            return False, "Name can only contain letters, spaces, hyphens, and apostrophes."
        if existing_names is None:
            existing_names = [e["name"] for e in self.employees]
        if name in existing_names:
            return False, f"'{name}' already exists."
        return True, ""

    def _on_canvas_mousewheel(self, event):
        """Scroll the preview canvas vertically."""
        delta = event.delta
        if sys.platform == "darwin":
            self.preview_canvas.yview_scroll(int(-1 * delta), "units")
        else:
            self.preview_canvas.yview_scroll(int(-1 * (delta / 120)), "units")

    def _on_canvas_shift_mousewheel(self, event):
        """Scroll the preview canvas horizontally."""
        delta = event.delta
        if sys.platform == "darwin":
            self.preview_canvas.xview_scroll(int(-1 * delta), "units")
        else:
            self.preview_canvas.xview_scroll(int(-1 * (delta / 120)), "units")

    def _bind_canvas_scroll(self):
        """Bind mouse wheel events to the canvas and all its children."""
        # Bind to canvas itself
        self.preview_canvas.bind("<MouseWheel>", self._on_canvas_mousewheel, add="+")
        self.preview_canvas.bind("<Shift-MouseWheel>", self._on_canvas_shift_mousewheel, add="+")

        # Also bind to the preview_frame and all its children
        if hasattr(self, 'preview_frame') and self.preview_frame.winfo_exists():
            self._bind_scroll_recursive(self.preview_frame)

    def _bind_scroll_recursive(self, widget):
        """Recursively bind mouse wheel to a widget and all its children."""
        try:
            widget.bind("<MouseWheel>", self._on_canvas_mousewheel, add="+")
            widget.bind("<Shift-MouseWheel>", self._on_canvas_shift_mousewheel, add="+")
            for child in widget.winfo_children():
                self._bind_scroll_recursive(child)
        except Exception:
            pass

    def check_manual_edit_violations(self, week_idx, day_idx, shift_idx, assigned_name):
        """
        Check if a manual edit violates any hard constraints.
        Returns a list of tuples: [(violation_type, conflicting_cell_key, short_message), ...]
        where violation_type is one of: 'overlap', 'type', 'timeoff'
        """
        if not self.generator or not assigned_name:
            return []

        violations = []

        # Get employee data
        emp_data = next((e for e in self.employees if e["name"] == assigned_name), None)
        if not emp_data:
            return []

        # 1. Check Type-Based Restrictions
        emp_type = emp_data.get("type", "")
        if emp_type == "late" and shift_idx == P1:
            violations.append(("type", None, "Wrong type"))
        elif emp_type == "early" and shift_idx == P4:
            violations.append(("type", None, "Wrong type"))

        # 2. Check Time-Off Violations
        # Use current time-off from UI to catch any changes made after generation
        engine_day = self.map_week_col_to_engine_day(week_idx, day_idx)
        if engine_day is not None and self.generator:
            # Get current time-off from UI
            current_timeoff_str = self.timeoff_entries.get(assigned_name, None)
            if current_timeoff_str:
                timeoff_str = current_timeoff_str.get().strip()
                if timeoff_str:
                    # Parse current time-off for this employee
                    month_num = self.generator.month
                    year = self.generator.year
                    max_day = calendar.monthrange(year, month_num)[1]

                    # Parse the time-off string
                    days_off = set()
                    for part in timeoff_str.split(","):
                        part = part.strip()
                        if not part:
                            continue
                        if "-" in part:
                            try:
                                start, end = part.split("-", 1)
                                start, end = int(start.strip()), int(end.strip())
                                if 1 <= start <= max_day and 1 <= end <= max_day and start <= end:
                                    days_off.update(range(start, end + 1))
                            except (ValueError, AttributeError):
                                pass
                        else:
                            try:
                                day_num = int(part)
                                if 1 <= day_num <= max_day:
                                    days_off.add(day_num)
                            except ValueError:
                                pass

                    # Check if this engine_day corresponds to a time-off day
                    if engine_day < len(self.generator.dates):
                        dt = self.generator.dates[engine_day]
                        if dt.day in days_off:
                            violations.append(("timeoff", None, "On time-off"))

        # 3. Check Overlapping Shifts (double duty same day)
        # Define which shifts overlap (time-based conflicts)
        overlaps = {
            P1: [FD1],
            P2: [FD1],
            P3: [FD2],
            P4: [FD2],
            FD1: [P1, P2],
            FD2: [P3, P4]
        }

        # Define same-category shifts (multiple phone or multiple FD shifts not allowed)
        same_category = {
            P1: [P2],
            P2: [P1],
            P3: [P4],
            P4: [P3],
            FD1: [FD2],
            FD2: [FD1]
        }

        # Check all other shifts on the same day for this employee
        for other_shift in overlaps.get(shift_idx, []):
            other_cell_key = (week_idx, day_idx, other_shift)
            if other_cell_key in self.edited_assignments:
                if self.edited_assignments[other_cell_key] == assigned_name:
                    violations.append(("overlap", other_cell_key, "Shift overlap"))

        # Check for same-category conflicts (multiple phone or FD shifts)
        for other_shift in same_category.get(shift_idx, []):
            other_cell_key = (week_idx, day_idx, other_shift)
            if other_cell_key in self.edited_assignments:
                if self.edited_assignments[other_cell_key] == assigned_name:
                    violations.append(("overlap", other_cell_key, "Multiple shifts"))

        # Also check if this person is already assigned to another shift on same day
        for s_idx in ALL_SHIFTS:
            if s_idx == shift_idx:
                continue
            check_key = (week_idx, day_idx, s_idx)
            if check_key in self.edited_assignments:
                if self.edited_assignments[check_key] == assigned_name:
                    # Check if these shifts overlap (time-based or same-category)
                    if s_idx in overlaps.get(shift_idx, []) or s_idx in same_category.get(shift_idx, []):
                        violations.append(("overlap", check_key, "Shift overlap"))

        return violations

    def map_week_col_to_engine_day(self, week_idx, col_idx):
        """Map preview grid coordinates to generator day index using the standalone helper."""
        return map_week_col_to_engine_day(self.generator, week_idx, col_idx)

    # ── Wizard ────────────────────────────────────────────────────────────────
    def show_wizard(self):
        self.clear_window()
        self.title("Schedule Builder — Team Setup")
        self.geometry("620x750")
        self.resizable(False, False)

        # Full-window gradient-like background
        bg = ctk.CTkFrame(self, fg_color=SURFACE_ALT, corner_radius=0)
        bg.place(relx=0, rely=0, relwidth=1, relheight=1)

        # Central card
        card = ctk.CTkFrame(bg, fg_color=SURFACE, corner_radius=RADIUS_LG,
                            border_width=1, border_color=BORDER)
        card.place(relx=0.5, rely=0.5, anchor="center", relwidth=0.88)

        # ── Header strip ──────────────────────────────────────────────────
        header = ctk.CTkFrame(card, fg_color=ACCENT, corner_radius=0,
                              height=6)
        header.pack(fill="x")
        # Give radius only on top by nesting
        header_top = ctk.CTkFrame(card, fg_color=ACCENT, corner_radius=RADIUS_LG,
                                  height=54)
        header_top.pack(fill="x")
        ctk.CTkLabel(
            header_top,
            text="  Team Setup",
            font=make_font(FS_XL, "bold"),
            text_color=TEXT_ON_ACCENT,
            anchor="w"
        ).place(relx=0.04, rely=0.5, anchor="w")

        body = ctk.CTkFrame(card, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=28, pady=20)
        body.grid_columnconfigure(0, weight=1)

        # Subtitle
        ctk.CTkLabel(
            body,
            text="Add at least 6 employees to generate schedules.",
            font=make_font(FS_SM),
            text_color=TEXT_MUTED,
            anchor="w"
        ).grid(row=0, column=0, sticky="w", pady=(0, 16))

        # ── Input row ─────────────────────────────────────────────────────
        input_card = ctk.CTkFrame(body, fg_color=SURFACE_ALT,
                                  corner_radius=RADIUS_MD,
                                  border_width=1, border_color=BORDER)
        input_card.grid(row=1, column=0, sticky="ew", pady=(0, 12))
        input_card.grid_columnconfigure(0, weight=1)

        section_label(input_card, "New Agent").grid(
            row=0, column=0, columnspan=2, sticky="w", padx=14, pady=(12, 6))

        inner = ctk.CTkFrame(input_card, fg_color="transparent")
        inner.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 12))
        inner.grid_columnconfigure(0, weight=1)

        self.wiz_name = ctk.CTkEntry(
            inner, placeholder_text="Full name…",
            font=make_font(FS_BASE),
            height=38, corner_radius=RADIUS_SM,
            border_color=BORDER
        )
        self.wiz_name.grid(row=0, column=0, sticky="ew", pady=(0, 8))

        self.wiz_email = ctk.CTkEntry(
            inner, placeholder_text="Email address…",
            font=make_font(FS_BASE),
            height=38, corner_radius=RADIUS_SM,
            border_color=BORDER
        )
        self.wiz_email.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        self.wiz_email.bind("<Return>", lambda e: self.wiz_add_employee())

        # Shift type selector
        type_row = ctk.CTkFrame(inner, fg_color="transparent")
        type_row.grid(row=2, column=0, sticky="w")
        self.wiz_type = ctk.StringVar(value="early")
        ctk.CTkLabel(type_row, text="Shift type:", font=make_font(FS_SM),
                     text_color=TEXT_MUTED).pack(side="left", padx=(0, 10))
        ctk.CTkRadioButton(type_row, text="Early  (7:30 – 4:00)",
                           variable=self.wiz_type, value="early",
                           font=make_font(FS_SM)).pack(side="left", padx=(0, 16))
        ctk.CTkRadioButton(type_row, text="Late  (8:30 – 5:00)",
                           variable=self.wiz_type, value="late",
                           font=make_font(FS_SM)).pack(side="left")

        ctk.CTkButton(
            input_card,
            text="+ Add Employee",
            command=self.wiz_add_employee,
            fg_color=ACCENT, hover_color=ACCENT_HOVER,
            text_color=TEXT_ON_ACCENT,
            font=make_font(FS_SM, "bold"),
            height=36, corner_radius=RADIUS_SM
        ).grid(row=2, column=0, sticky="ew", padx=14, pady=(0, 14))

        # ── Employee list ─────────────────────────────────────────────────
        section_label(body, "Roster").grid(
            row=2, column=0, sticky="w", pady=(4, 6))

        list_frame = ctk.CTkFrame(body, fg_color=SURFACE_ALT,
                                  corner_radius=RADIUS_MD,
                                  border_width=1, border_color=BORDER)
        list_frame.grid(row=3, column=0, sticky="ew")

        self.wiz_list_inner = ctk.CTkScrollableFrame(
            list_frame, height=190, fg_color="transparent",
            label_text="",
            scrollbar_fg_color="transparent",
            scrollbar_button_color=SCROLLBAR,
            scrollbar_button_hover_color=ACCENT
        )
        self.wiz_list_inner.pack(fill="both", expand=True, padx=4, pady=4)
        self.wiz_list_inner.grid_columnconfigure(0, weight=1)

        self.wiz_employees = []
        self._wiz_row_count = 0

        # Count badge
        self.wiz_count_label = ctk.CTkLabel(
            body, text="0 / 6 minimum",
            font=make_font(FS_XS), text_color=TEXT_MUTED
        )
        self.wiz_count_label.grid(row=4, column=0, sticky="e", pady=(4, 0))

        # ── Footer ────────────────────────────────────────────────────────
        divider_grid(body, row=5, pady=(14, 10))

        ctk.CTkButton(
            body,
            text="Save & Launch  →",
            command=self.wiz_finish,
            fg_color=ACCENT, hover_color=ACCENT_HOVER,
            text_color=TEXT_ON_ACCENT,
            font=make_font(FS_LG, "bold"),
            height=42, corner_radius=RADIUS_SM
        ).grid(row=6, column=0, sticky="ew", pady=(0, 4))

    def _wiz_refresh_count(self):
        n = len(self.wiz_employees)
        color = TEXT_MUTED if n >= 6 else RED_TEXT
        self.wiz_count_label.configure(
            text=f"{n} employee{'s' if n != 1 else ''} added  (min 6)",
            text_color=color
        )

    def wiz_add_employee(self):
        name = self.wiz_name.get().strip()
        email = self.wiz_email.get().strip()
        e_type = self.wiz_type.get()

        wiz_names = [e["name"] for e in self.wiz_employees]
        is_valid, error_msg = self.validate_employee_name(name, existing_names=wiz_names)
        if not is_valid:
            messagebox.showwarning("Input Error", error_msg)
            return

        # Basic email validation (if provided)
        if email and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
            messagebox.showwarning("Input Error", "Please enter a valid email address.")
            return

        emp = {"name": name, "type": e_type, "email": email}
        self.wiz_employees.append(emp)
        self.wiz_name.delete(0, "end")
        self.wiz_email.delete(0, "end")

        # Add row to scrollable list
        self._wiz_add_row(emp)
        self._wiz_refresh_count()

    def _wiz_add_row(self, emp):
        """Add a single employee row to the wizard roster list."""
        name = emp["name"]
        e_type = emp["type"]

        row_frame = ctk.CTkFrame(
            self.wiz_list_inner,
            fg_color=SURFACE,
            corner_radius=RADIUS_SM,
            border_width=1, border_color=BORDER
        )
        row_frame.pack(fill="x", pady=3, padx=4)

        badge_color = BLUE_BG if e_type == "early" else ORANGE_BG
        badge_text_color = BLUE_TEXT if e_type == "early" else ORANGE_TEXT
        badge_label = "Early" if e_type == "early" else "Late"

        ctk.CTkLabel(
            row_frame, text=name,
            font=make_font(FS_SM, "bold"),
            text_color=TEXT_PRIMARY, anchor="w"
        ).pack(side="left", padx=12, pady=8)

        def remove():
            if emp in self.wiz_employees:
                self.wiz_employees.remove(emp)
            row_frame.destroy()
            self._wiz_refresh_count()

        ctk.CTkButton(
            row_frame, text="x", width=24, height=24,
            corner_radius=12,
            fg_color="transparent", hover_color=RED_BG,
            text_color=RED_TEXT,
            font=make_font(FS_XS, "bold"),
            command=remove
        ).pack(side="right", padx=(4, 8))

        ctk.CTkLabel(
            row_frame, text=badge_label,
            font=make_font(FS_XS, "bold"),
            fg_color=badge_color, text_color=badge_text_color,
            corner_radius=20, width=48, height=20
        ).pack(side="right", padx=(0, 4))

    def wiz_finish(self):
        if len(self.wiz_employees) < 6:
            if not messagebox.askyesno(
                "Too Few Employees",
                f"You only have {len(self.wiz_employees)} employee(s).\n"
                "Coverage requires at least 6 (6 shifts/day).\n\n"
                "Continue anyway?"
            ):
                return
        self.config_manager.save_config(self.wiz_employees)
        self.employees = self.wiz_employees
        self.resizable(True, True)
        
        # Remove focus from any active entry widgets to prevent TclError
        self.focus_set()
        
        # Defer the transition so CustomTkinter button/entry events can finish resolving
        self.after(50, self.show_dashboard)

    # ── Dashboard ─────────────────────────────────────────────────────────────
    def show_dashboard(self):
        self.clear_window()
        self.title("Schedule Builder")
        self.geometry("1180x820")
        self.minsize(900, 680)
        self.resizable(True, True)

        self.grid_columnconfigure(0, weight=0)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # ── Sidebar ───────────────────────────────────────────────────────
        self.sidebar = ctk.CTkFrame(
            self, width=414,  # 20% wider than 345
            fg_color=SIDEBAR_BG,
            corner_radius=0,
            border_width=0
        )
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.grid_propagate(False)
        self.sidebar.grid_rowconfigure(1, weight=1)
        self.sidebar.grid_columnconfigure(0, weight=1)

        # Brand header strip at top of sidebar
        brand = ctk.CTkFrame(self.sidebar, fg_color=ACCENT, corner_radius=0, height=56)
        brand.grid(row=0, column=0, sticky="ew")
        brand.grid_propagate(False)
        ctk.CTkLabel(
            brand, text="Schedule Builder",
            font=make_font(FS_LG, "bold"),
            text_color=TEXT_ON_ACCENT
        ).place(relx=0.5, rely=0.5, anchor="center")

        # Scrollable sidebar body
        self.sidebar_scroll = ctk.CTkScrollableFrame(
            self.sidebar, fg_color="transparent",
            scrollbar_fg_color="transparent",
            scrollbar_button_color=SCROLLBAR,
            scrollbar_button_hover_color=ACCENT
        )
        self.sidebar_scroll.grid(row=1, column=0, sticky="nsew")
        self.sidebar_scroll.grid_columnconfigure(0, weight=1)

        self._build_sidebar_controls(self.sidebar_scroll)

        # Bottom actions (fixed, not scrollable)
        bottom = ctk.CTkFrame(self.sidebar, fg_color=SIDEBAR_BG,
                              corner_radius=0)
        bottom.grid(row=2, column=0, sticky="ew")
        sep = ctk.CTkFrame(bottom, height=1, fg_color=BORDER, corner_radius=0)
        sep.pack(fill="x")

        # Single Settings button
        btn_settings = ctk.CTkButton(
            bottom, text="Settings",
            command=self.show_settings_menu,
            fg_color=SURFACE, hover_color=SURFACE_ALT,
            text_color=TEXT_PRIMARY,
            border_width=1, border_color=BORDER,
            font=make_font(FS_SM), height=34,
            corner_radius=RADIUS_SM
        )
        btn_settings.pack(fill="x", padx=16, pady=12)

        # ── Main area ─────────────────────────────────────────────────────
        self.main_frame = ctk.CTkFrame(self, fg_color=SURFACE_ALT, corner_radius=0)
        self.main_frame.grid(row=0, column=1, sticky="nsew")
        self.main_frame.grid_rowconfigure(1, weight=1)
        self.main_frame.grid_columnconfigure(0, weight=1)
        self.main_frame.grid_columnconfigure(1, weight=0)  # side panel col
        self.main_frame.grid_columnconfigure(2, weight=0)  # vsb col

        # Top bar
        topbar = ctk.CTkFrame(self.main_frame, fg_color=SURFACE,
                              corner_radius=0, height=56,
                              border_width=1, border_color=BORDER)
        topbar.grid(row=0, column=0, columnspan=3, sticky="ew")
        topbar.grid_propagate(False)

        ctk.CTkLabel(
            topbar, text="Schedule Preview",
            font=make_font(FS_LG, "bold"),
            text_color=TEXT_PRIMARY
        ).place(relx=0.02, rely=0.5, anchor="w")

        # Export / action buttons — right-aligned row
        btn_row = ctk.CTkFrame(topbar, fg_color="transparent")
        btn_row.place(relx=0.98, rely=0.5, anchor="e")

        # Pack in REVERSE order (side="right" means first packed = rightmost)
        # Desired order L→R: Generate, Excel, Calendar, Email, Save & Exit

        # Save & Exit button (rightmost)
        self.btn_save_session = ctk.CTkButton(
            btn_row, text="Save & Exit",
            command=self.save_and_exit, state="disabled",
            fg_color=SURFACE_ALT, hover_color=BORDER,
            text_color=TEXT_PRIMARY,
            border_width=1, border_color=BORDER,
            font=make_font(FS_SM, "bold"),
            height=34, corner_radius=RADIUS_SM,
            width=120
        )
        self.btn_save_session.pack(side="right", padx=(6, 0))

        # Email Schedule
        self.btn_email = ctk.CTkButton(
            btn_row, text="Email Schedule",
            command=self.email_schedule_dialog, state="disabled",
            fg_color=SURFACE_ALT, hover_color=BORDER,
            text_color=TEXT_PRIMARY,
            border_width=1, border_color=BORDER,
            font=make_font(FS_SM, "bold"),
            height=34, corner_radius=RADIUS_SM,
            width=130
        )
        self.btn_email.pack(side="right", padx=(6, 0))

        # Export Calendar
        self.btn_ics = ctk.CTkButton(
            btn_row, text="Export Calendar",
            command=self.export_calendar_dialog, state="disabled",
            fg_color=SURFACE_ALT, hover_color=BORDER,
            text_color=TEXT_PRIMARY,
            border_width=1, border_color=BORDER,
            font=make_font(FS_SM, "bold"),
            height=34, corner_radius=RADIUS_SM,
            width=130
        )
        self.btn_ics.pack(side="right", padx=(6, 0))

        # Export to Excel
        self.btn_save = ctk.CTkButton(
            btn_row, text="Export to Excel",
            command=self.save_excel, state="disabled",
            fg_color=SURFACE_ALT, hover_color=BORDER,
            text_color=TEXT_PRIMARY,
            border_width=1, border_color=BORDER,
            font=make_font(FS_SM, "bold"),
            height=34, corner_radius=RADIUS_SM,
            width=130
        )
        self.btn_save.pack(side="right", padx=(6, 0))

        # Generate Schedule (leftmost of action buttons)
        self.btn_generate = ctk.CTkButton(
            btn_row, text="Generate Schedule",
            command=self.run_scheduler, state="normal",
            fg_color=ACCENT, hover_color=ACCENT_HOVER,
            text_color=TEXT_ON_ACCENT,
            font=make_font(FS_SM, "bold"),
            height=34, corner_radius=RADIUS_SM,
            width=140
        )
        self.btn_generate.pack(side="right", padx=(6, 0))

        # Tooltips for the export/action buttons
        ToolTip(self.btn_save,
                "Export the schedule to an Excel (.xlsx) file\n"
                "with colour-coded shifts and a fairness audit.")
        ToolTip(self.btn_ics,
                "Export .ics calendar files that agents can\n"
                "import into Outlook, Google Calendar, etc.")
        ToolTip(self.btn_email,
                "Open an Outlook email draft with each agent's\n"
                "calendar file attached — review before sending.")
        ToolTip(self.btn_save_session,
                "Save the current schedule and exit the application.\n"
                "Resume your work when you open the app again.")

        # Canvas + scrollbars
        self.preview_canvas = ctk.CTkCanvas(
            self.main_frame, highlightthickness=0, borderwidth=0,
            bg=self._surface_alt_hex()
        )
        self.preview_canvas.grid(row=1, column=0, sticky="nsew")

        # ── Side panel (fairness audit + action panel) ────────────────────
        self.side_panel = ctk.CTkFrame(
            self.main_frame, fg_color=SURFACE_ALT,
            corner_radius=0, width=360
        )
        self.side_panel.grid(row=1, column=1, rowspan=2, sticky="nsew")
        self.side_panel.grid_propagate(False)
        self.side_panel.pack_propagate(False)  # Prevent shrinking based on packed children

        # Fairness audit frame (top of side panel)
        self.fairness_frame = ctk.CTkFrame(
            self.side_panel, fg_color="transparent"
        )
        self.fairness_frame.pack(fill="x", padx=8, pady=(8, 0))

        # Action panel (bottom of side panel, hidden by default)
        self.action_panel = ctk.CTkFrame(
            self.side_panel, fg_color=SURFACE,
            corner_radius=RADIUS_MD,
            border_width=1, border_color=BORDER
        )
        # Initially hidden — shown via _show_action_panel()

        self.preview_vsb = ctk.CTkScrollbar(
            self.main_frame, orientation="vertical",
            command=self.preview_canvas.yview,
            button_color=SCROLLBAR,
            button_hover_color=ACCENT
        )
        self.preview_vsb.grid(row=1, column=2, sticky="ns")

        self.preview_hsb = ctk.CTkScrollbar(
            self.main_frame, orientation="horizontal",
            command=self.preview_canvas.xview,
            button_color=SCROLLBAR,
            button_hover_color=ACCENT
        )
        self.preview_hsb.grid(row=2, column=0, sticky="ew")

        self.preview_canvas.configure(
            yscrollcommand=self.preview_vsb.set,
            xscrollcommand=self.preview_hsb.set
        )

        self.preview_frame = ctk.CTkFrame(
            self.preview_canvas, fg_color=SURFACE_ALT
        )
        self.preview_canvas.create_window((0, 0), window=self.preview_frame, anchor="nw")
        self.preview_frame.bind(
            "<Configure>",
            lambda e: self.preview_canvas.configure(
                scrollregion=self.preview_canvas.bbox("all")
            )
        )

        # Bind mouse wheel scrolling to the canvas
        self._bind_canvas_scroll()

        # Keep the raw Tkinter canvas background in sync with the OS appearance
        ctk.AppearanceModeTracker.add(self._on_appearance_change)

        self._show_empty_state()

        # Check for a saved session and offer to resume
        self.after(200, self._offer_resume)

    def _surface_alt_hex(self):
        """Return hex for canvas bg depending on the current appearance mode."""
        mode = ctk.get_appearance_mode()
        return SURFACE_ALT[0] if mode == "Light" else SURFACE_ALT[1]

    def _on_appearance_change(self, mode_string):
        """Called by CustomTkinter when the OS appearance mode changes."""
        # Update the raw Tkinter canvas background (it doesn't auto-update)
        if hasattr(self, "preview_canvas") and self.preview_canvas.winfo_exists():
            self.preview_canvas.configure(bg=self._surface_alt_hex())
        # If a schedule is displayed, re-render it so resolved colours are fresh
        if self.generator is not None and hasattr(self, "preview_frame"):
            self.render_preview(self.generator)

    # ── Action Panel (inline replacement for dialog windows) ──────────────

    def _show_action_panel(self, title, build_content_fn, back_command=None):
        """Show the inline action panel with content built by build_content_fn.

        Args:
            title: Panel title
            build_content_fn: Function that builds panel content
            back_command: Optional function to call when back button is pressed
        """
        for w in self.action_panel.winfo_children():
            w.destroy()

        # Header with title and close button (only if title is provided)
        if title:
            header = ctk.CTkFrame(self.action_panel, fg_color=ACCENT,
                                  height=40, corner_radius=0)
            header.pack(fill="x")
            header.pack_propagate(False)
            ctk.CTkLabel(
                header, text=title,
                font=make_font(FS_SM, "bold"),
                text_color=TEXT_ON_ACCENT
            ).place(relx=0.04, rely=0.5, anchor="w")

            # Close button (X)
            ctk.CTkButton(
                header, text="X", width=26, height=26, corner_radius=13,
                fg_color="transparent", hover_color=ACCENT_HOVER,
                text_color=TEXT_ON_ACCENT,
                font=make_font(FS_SM, "bold"),
                command=self._hide_action_panel
            ).place(relx=0.96, rely=0.5, anchor="e")

            # Back button (if back_command is provided)
            if back_command:
                ctk.CTkButton(
                    header, text="← Back", width=60, height=26, corner_radius=13,
                    fg_color="transparent", hover_color=ACCENT_HOVER,
                    text_color=TEXT_ON_ACCENT,
                    font=make_font(FS_XS, "bold"),
                    command=back_command
                ).place(relx=0.85, rely=0.5, anchor="e")

        # Scrollable content area
        content = ctk.CTkScrollableFrame(
            self.action_panel, fg_color="transparent",
            scrollbar_fg_color="transparent",
            scrollbar_button_color=SCROLLBAR,
            scrollbar_button_hover_color=ACCENT
        )
        content.pack(fill="both", expand=True, padx=8, pady=8)

        build_content_fn(content)

        # Show the panel
        self.action_panel.pack(fill="both", expand=True, padx=8, pady=(8, 8))

    def _hide_action_panel(self):
        """Hide the inline action panel."""
        self.action_panel.pack_forget()

    # ── Fairness Audit (rendered in side panel) ───────────────────────────

    def _get_stats_with_edits(self, generator):
        """Calculate stats including manual edits."""
        stats = {}
        staff_names = [e["name"] for e in self.employees]

        # Initialize counters
        for name in staff_names:
            stats[name] = {"name": name, "phone": 0, "fd": 0, "total": 0}

        # Count shifts from the schedule, applying manual edits
        weeks = generator.get_weekly_matrix()
        for week_idx, week_data in enumerate(weeks):
            for day_idx in range(5):  # Mon-Fri
                for shift_idx in ALL_SHIFTS:
                    # Check for manual edit first
                    cell_key = (week_idx, day_idx, shift_idx)
                    if cell_key in self.edited_assignments:
                        worker = self.edited_assignments[cell_key]
                    else:
                        worker = week_data["matrix"][shift_idx][day_idx]

                    if worker and worker in stats:
                        if shift_idx in PHONE_SHIFTS:
                            stats[worker]["phone"] += 1
                        elif shift_idx in FD_SHIFTS:
                            stats[worker]["fd"] += 1
                        stats[worker]["total"] += 1

        return list(stats.values())

    def _render_fairness_panel(self, generator):
        """Render the fairness audit table in the side panel."""
        for w in self.fairness_frame.winfo_children():
            w.destroy()

        mode = ctk.get_appearance_mode()
        is_dark = (mode == "Dark")

        def resolve(token):
            if isinstance(token, tuple):
                return token[1] if is_dark else token[0]
            return token

        r_header = ACCENT
        r_text = resolve(TEXT_PRIMARY)
        r_surface = resolve(SURFACE)

        audit_card = ctk.CTkFrame(
            self.fairness_frame,
            fg_color=r_surface,
            corner_radius=RADIUS_MD,
            border_width=1,
            border_color=resolve(BORDER)
        )
        audit_card.pack(fill="x")

        audit_header = ctk.CTkFrame(audit_card, fg_color="transparent", height=40)
        audit_header.pack(fill="x", padx=12, pady=(10, 4))
        ctk.CTkLabel(
            audit_header,
            text="Fairness Audit",
            font=make_font(FS_LG, "bold"),
            text_color=TEXT_PRIMARY
        ).pack(side="left")

        audit_table = ctk.CTkFrame(audit_card, fg_color="transparent")
        audit_table.pack(fill="both", expand=True, padx=10, pady=(0, 12))

        headers = ["Employee", "Phone", "FD", "Total"]
        col_w = [120, 60, 60, 60]
        for col_i, (h, w) in enumerate(zip(headers, col_w)):
            ctk.CTkLabel(
                audit_table,
                text=h,
                width=w, height=34,
                fg_color=r_header,
                text_color=TEXT_ON_ACCENT,
                font=make_font(FS_XS, "bold"),
                corner_radius=RADIUS_SM,
                anchor="center"
            ).grid(row=0, column=col_i, padx=2, pady=2, sticky="nsew")

        stats = self._get_stats_with_edits(generator)
        for row_i, stat in enumerate(stats):
            row_bg = resolve(SURFACE_ALT) if row_i % 2 == 1 else r_surface
            values = [stat["name"], str(stat["phone"]),
                      str(stat["fd"]), str(stat["total"])]
            for col_i, (val, w) in enumerate(zip(values, col_w)):
                is_total = col_i == 3
                ctk.CTkLabel(
                    audit_table,
                    text=val,
                    width=w, height=30,
                    fg_color=row_bg,
                    text_color=r_text,
                    font=make_font(FS_XS, "bold" if is_total else "normal"),
                    corner_radius=RADIUS_SM,
                    anchor="w" if col_i == 0 else "center",
                    padx=8 if col_i == 0 else 0
                ).grid(row=row_i + 1, column=col_i,
                       padx=2, pady=1, sticky="nsew")

    def _build_sidebar_controls(self, parent):
        """Populate the scrollable sidebar with all control sections."""
        pad_x = 16

        # ── Schedule Configuration ─────────────────────────────────────
        section_label(parent, "Schedule").grid(
            row=0, column=0, sticky="w", padx=pad_x, pady=(18, 6)
        )

        config_card = ctk.CTkFrame(
            parent, fg_color=SURFACE,
            corner_radius=RADIUS_MD,
            border_width=1, border_color=BORDER
        )
        config_card.grid(row=1, column=0, sticky="ew", padx=pad_x)
        config_card.grid_columnconfigure(0, weight=1)

        inner_pad = 14

        # Month
        ctk.CTkLabel(
            config_card, text="Month",
            font=make_font(FS_XS, "bold"), text_color=TEXT_MUTED, anchor="w"
        ).grid(row=0, column=0, sticky="w", padx=inner_pad, pady=(12, 2))
        self.month_var = ctk.StringVar(value=datetime.now().strftime("%B"))
        self.opt_month = ctk.CTkOptionMenu(
            config_card,
            values=list(calendar.month_name)[1:],
            variable=self.month_var,
            font=make_font(FS_SM),
            fg_color=SURFACE_ALT,
            button_color=ACCENT, button_hover_color=ACCENT_HOVER,
            dropdown_fg_color=SURFACE,
            corner_radius=RADIUS_SM,
            height=36
        )
        self.opt_month.grid(row=1, column=0, sticky="ew", padx=inner_pad, pady=(0, 8))

        # Year
        ctk.CTkLabel(
            config_card, text="Year",
            font=make_font(FS_XS, "bold"), text_color=TEXT_MUTED, anchor="w"
        ).grid(row=2, column=0, sticky="w", padx=inner_pad, pady=(0, 2))
        current_year = datetime.now().year
        years = [str(y) for y in range(current_year, 2051)]
        self.year_var = ctk.StringVar(value=str(current_year))
        self.opt_year = ctk.CTkOptionMenu(
            config_card,
            values=years,
            variable=self.year_var,
            font=make_font(FS_SM),
            fg_color=SURFACE_ALT,
            button_color=ACCENT, button_hover_color=ACCENT_HOVER,
            dropdown_fg_color=SURFACE,
            corner_radius=RADIUS_SM,
            height=36
        )
        self.opt_year.grid(row=3, column=0, sticky="ew", padx=inner_pad, pady=(0, 8))

        # Holidays
        ctk.CTkLabel(
            config_card, text="Holidays",
            font=make_font(FS_XS, "bold"), text_color=TEXT_MUTED, anchor="w"
        ).grid(row=4, column=0, sticky="w", padx=inner_pad, pady=(0, 2))
        self.entry_holidays = ctk.CTkEntry(
            config_card,
            placeholder_text="e.g.  1, 25-30",
            font=make_font(FS_SM),
            height=36, corner_radius=RADIUS_SM,
            border_color=BORDER
        )
        self.entry_holidays.grid(row=5, column=0, sticky="ew",
                                 padx=inner_pad, pady=(0, 14))

        # ── Time Off Requests ──────────────────────────────────────────
        section_label(parent, "Time Off Requests").grid(
            row=2, column=0, sticky="w", padx=pad_x, pady=(18, 6)
        )

        timeoff_card = ctk.CTkFrame(
            parent, fg_color=SURFACE,
            corner_radius=RADIUS_MD,
            border_width=1, border_color=BORDER
        )
        timeoff_card.grid(row=3, column=0, sticky="ew", padx=pad_x)
        timeoff_card.grid_columnconfigure(0, weight=1)

        hint = ctk.CTkLabel(
            timeoff_card,
            text="Enter day numbers or ranges, e.g. 1, 5, 10-15, 20",
            font=make_font(FS_XS), text_color=TEXT_MUTED, anchor="w"
        )
        hint.grid(row=0, column=0, sticky="w", padx=inner_pad, pady=(10, 6))

        self.frame_timeoff = ctk.CTkScrollableFrame(
            timeoff_card,
            height=220,
            fg_color="transparent",
            label_text="",
            scrollbar_fg_color="transparent",
            scrollbar_button_color=SCROLLBAR,
            scrollbar_button_hover_color=ACCENT
        )
        self.frame_timeoff.grid(row=1, column=0, sticky="ew",
                                padx=6, pady=(0, 8))
        self.frame_timeoff.grid_columnconfigure(0, weight=1)

        self.setup_timeoff_ui()

        # Clear time-off button
        ctk.CTkButton(
            parent,
            text="Clear All Time Off",
            command=self.clear_timeoff,
            fg_color="transparent",
            hover_color=SURFACE_ALT,
            text_color=TEXT_MUTED,
            border_width=1, border_color=BORDER,
            font=make_font(FS_XS),
            height=28, corner_radius=RADIUS_SM
        ).grid(row=4, column=0, sticky="ew", padx=pad_x, pady=(0, 18))

        # ── Advanced Options ────────────────────────────────────────────
        section_label(parent, "Advanced Scheduling").grid(
            row=5, column=0, sticky="w", padx=pad_x, pady=(8, 6)
        )

        advanced_card = ctk.CTkFrame(
            parent, fg_color=SURFACE,
            corner_radius=RADIUS_MD,
            border_width=1, border_color=BORDER
        )
        advanced_card.grid(row=6, column=0, sticky="ew", padx=pad_x)
        advanced_card.grid_columnconfigure(0, weight=1)

        # Full calendar weeks toggle
        self.full_weeks_var = ctk.BooleanVar(value=False)
        full_weeks_check = ctk.CTkCheckBox(
            advanced_card,
            text="Complete partial weeks into next month",
            variable=self.full_weeks_var,
            font=make_font(FS_SM),
            fg_color=ACCENT,
            hover_color=ACCENT_HOVER
        )
        full_weeks_check.grid(row=0, column=0, sticky="w", padx=inner_pad, pady=(12, 4))
        ToolTip(full_weeks_check,
                "When the month ends mid-week, extend the schedule\n"
                "to complete the full week (ending on Friday).\n"
                "Example: If month ends on Tuesday, schedule continues through Friday.")

        # Start day field
        start_day_frame = ctk.CTkFrame(advanced_card, fg_color="transparent")
        start_day_frame.grid(row=1, column=0, sticky="ew", padx=inner_pad, pady=(4, 12))

        start_day_label = ctk.CTkLabel(
            start_day_frame,
            text="Start schedule from day:",
            font=make_font(FS_SM),
            text_color=TEXT_PRIMARY
        )
        start_day_label.pack(side="left", padx=(0, 8))

        self.start_day_entry = ctk.CTkEntry(
            start_day_frame,
            placeholder_text="1",
            width=70,
            font=make_font(FS_SM),
            height=32
        )
        self.start_day_entry.pack(side="left")

        ToolTip(start_day_label,
                "Generate schedule starting from a specific day.\n"
                "Example: Enter '15' to schedule only from the 15th onwards.\n"
                "Useful for partial month schedules or mid-month changes.")

    def setup_timeoff_ui(self):
        """Populate the scrollable frame with one row per employee."""
        for widget in self.frame_timeoff.winfo_children():
            widget.destroy()
        self.timeoff_entries = {}
        self.employees = self.config_manager.get_employees()

        # Dynamically adjust height based on number of employees
        # Each row is ~40px (badge/entry height + pady)
        # Grow up to 10 agents, then start scrolling
        num_employees = len(self.employees)
        row_height = 40
        max_agents_before_scroll = 10
        min_height = 60  # Minimum height for empty state or few agents

        if num_employees == 0:
            new_height = min_height
        elif num_employees <= max_agents_before_scroll:
            new_height = max(min_height, num_employees * row_height + 20)  # +20 for padding
        else:
            new_height = max_agents_before_scroll * row_height + 20

        self.frame_timeoff.configure(height=new_height)

        if not self.employees:
            ctk.CTkLabel(
                self.frame_timeoff, text="No employees found.",
                font=make_font(FS_SM), text_color=TEXT_MUTED
            ).pack(pady=10)
            return

        for emp in self.employees:
            name = emp["name"]
            etype = emp.get("type", "early")

            row = ctk.CTkFrame(self.frame_timeoff, fg_color="transparent")
            row.pack(fill="x", pady=3)
            row.grid_columnconfigure(1, weight=1)

            # Badge
            badge_color = BLUE_BG if etype == "early" else ORANGE_BG
            badge_text_color = BLUE_TEXT if etype == "early" else ORANGE_TEXT
            ctk.CTkLabel(
                row,
                text=name[0].upper(),
                width=28, height=28,
                fg_color=badge_color,
                text_color=badge_text_color,
                corner_radius=14,
                font=make_font(FS_XS, "bold")
            ).grid(row=0, column=0, padx=(4, 6))

            ctk.CTkLabel(
                row, text=name,
                font=make_font(FS_SM),
                text_color=TEXT_PRIMARY,
                anchor="w",
                width=80
            ).grid(row=0, column=1, sticky="w")

            entry = ctk.CTkEntry(
                row,
                placeholder_text="days…",
                font=make_font(FS_SM),
                height=30, corner_radius=RADIUS_SM,
                border_color=BORDER,
                width=80
            )
            entry.grid(row=0, column=2, padx=(4, 4))
            self.timeoff_entries[name] = entry

    def clear_timeoff(self):
        for entry in self.timeoff_entries.values():
            entry.delete(0, "end")

    # ── Empty state ───────────────────────────────────────────────────────────
    # ── Session Save / Resume ────────────────────────────────────────────────

    def save_session(self):
        """Serialize sidebar inputs + schedule data to disk."""
        if not self.generator:
            return

        # Collect time-off entry values
        timeoff_raw = {}
        for name, entry in self.timeoff_entries.items():
            val = entry.get().strip()
            if val:
                timeoff_raw[name] = val

        # Convert tuple-keyed edited_assignments to JSON-safe string keys
        edits_json = {}
        for (wi, di, si), worker in self.edited_assignments.items():
            edits_json[f"{wi},{di},{si}"] = worker

        # Save the actual schedule assignments (day, shift) -> employee
        schedule_data = {}
        staff_names = [e["name"] for e in self.employees]
        for d in range(self.generator.num_days):
            for s in ALL_SHIFTS:
                worker = ""
                for emp in staff_names:
                    if self.generator.solver.Value(self.generator.shifts[(emp, d, s)]):
                        worker = emp
                        break
                if worker:  # Only save non-empty assignments
                    schedule_data[f"{d},{s}"] = worker

        data = {
            "month": self.month_var.get(),
            "year": self.year_var.get(),
            "holidays": self.entry_holidays.get().strip(),
            "timeoff": timeoff_raw,
            "full_weeks": self.full_weeks_var.get(),
            "start_day": self.start_day_entry.get().strip(),
            "edited_assignments": edits_json,
            "schedule": schedule_data,
        }

        try:
            self.config_manager.save_session(data)
            def build(f):
                ctk.CTkLabel(
                    f, text="Session saved! Next time you open the app, "
                    "you'll be offered to resume from this point.",
                    font=make_font(FS_SM), text_color=TEXT_PRIMARY,
                    wraplength=320, justify="left"
                ).pack(anchor="w")
            self._show_action_panel("Session Saved", build)
        except Exception as e:
            def build(f, err=str(e)):
                ctk.CTkLabel(
                    f, text=f"Failed to save session:\n{err}",
                    font=make_font(FS_SM), text_color=RED_TEXT,
                    wraplength=320, justify="left"
                ).pack(anchor="w")
            self._show_action_panel("Save Error", build)

    def save_and_exit(self):
        """Save session silently and exit the application after showing confirmation."""
        if not self.generator:
            return

        # Collect time-off entry values
        timeoff_raw = {}
        for name, entry in self.timeoff_entries.items():
            val = entry.get().strip()
            if val:
                timeoff_raw[name] = val

        # Convert tuple-keyed edited_assignments to JSON-safe string keys
        edits_json = {}
        for (wi, di, si), worker in self.edited_assignments.items():
            edits_json[f"{wi},{di},{si}"] = worker

        # Save the actual schedule assignments (day, shift) -> employee
        schedule_data = {}
        staff_names = [e["name"] for e in self.employees]
        for d in range(self.generator.num_days):
            for s in ALL_SHIFTS:
                worker = ""
                for emp in staff_names:
                    if self.generator.solver.Value(self.generator.shifts[(emp, d, s)]):
                        worker = emp
                        break
                if worker:  # Only save non-empty assignments
                    schedule_data[f"{d},{s}"] = worker

        data = {
            "month": self.month_var.get(),
            "year": self.year_var.get(),
            "holidays": self.entry_holidays.get().strip(),
            "timeoff": timeoff_raw,
            "full_weeks": self.full_weeks_var.get(),
            "start_day": self.start_day_entry.get().strip(),
            "edited_assignments": edits_json,
            "schedule": schedule_data,
        }

        try:
            # Save silently in background
            self.config_manager.save_session(data)

            # Show "Saved, Exiting..." message
            def build(f):
                ctk.CTkLabel(
                    f, text="Saved, Exiting...",
                    font=make_font(FS_LG, "bold"),
                    text_color=TEXT_PRIMARY
                ).pack(anchor="center", pady=40)

            self._show_action_panel("", build)  # No title for cleaner look

            # Close application after 2 seconds
            self.after(2000, self.quit)

        except Exception as e:
            # If save fails, show error and don't exit
            def build(f, err=str(e)):
                ctk.CTkLabel(
                    f, text=f"Failed to save session:\n{err}",
                    font=make_font(FS_SM), text_color=RED_TEXT,
                    wraplength=320, justify="left"
                ).pack(anchor="w")
            self._show_action_panel("Save Error", build)

    def _offer_resume(self):
        """Check for a saved session and show a resume prompt in the action panel."""
        session = self.config_manager.load_session()
        if not session:
            return

        month = session.get("month", "?")
        year = session.get("year", "?")
        n_edits = len(session.get("edited_assignments", {}))
        edit_note = f" with {n_edits} manual edit(s)" if n_edits else ""

        def build(f):
            ctk.CTkLabel(
                f, text=f"Found a saved session for {month} {year}{edit_note}.",
                font=make_font(FS_SM), text_color=TEXT_PRIMARY,
                wraplength=320, justify="left"
            ).pack(anchor="w", pady=(0, 4))
            ctk.CTkLabel(
                f, text="Would you like to resume where you left off? "
                "This will restore your saved schedule exactly as it was.",
                font=make_font(FS_SM), text_color=TEXT_MUTED,
                wraplength=320, justify="left"
            ).pack(anchor="w", pady=(0, 14))

            ctk.CTkButton(
                f, text="Resume Session", command=lambda: self._do_resume(session),
                fg_color=ACCENT, hover_color=ACCENT_HOVER,
                text_color=TEXT_ON_ACCENT,
                font=make_font(FS_SM, "bold"),
                height=34, corner_radius=RADIUS_SM
            ).pack(fill="x", pady=(0, 6))

            def discard():
                self.config_manager.delete_session()
                self._hide_action_panel()

            ctk.CTkButton(
                f, text="Discard Saved Session", command=discard,
                fg_color=SURFACE_ALT, hover_color=BORDER,
                text_color=TEXT_PRIMARY,
                border_width=1, border_color=BORDER,
                font=make_font(FS_SM, "bold"),
                height=34, corner_radius=RADIUS_SM
            ).pack(fill="x")

        self._show_action_panel("Resume Session?", build)

    def _do_resume(self, session):
        """Restore sidebar inputs from saved session and re-run the solver."""
        self._hide_action_panel()

        # Restore sidebar inputs
        self.month_var.set(session.get("month", self.month_var.get()))
        self.year_var.set(session.get("year", self.year_var.get()))

        # Holidays
        holidays_val = session.get("holidays", "")
        self.entry_holidays.delete(0, "end")
        if holidays_val:
            self.entry_holidays.insert(0, holidays_val)

        # Time-off entries
        saved_timeoff = session.get("timeoff", {})
        for name, entry in self.timeoff_entries.items():
            entry.delete(0, "end")
            if name in saved_timeoff:
                entry.insert(0, saved_timeoff[name])

        # Advanced options
        self.full_weeks_var.set(session.get("full_weeks", False))
        start_day_val = session.get("start_day", "")
        self.start_day_entry.delete(0, "end")
        if start_day_val:
            self.start_day_entry.insert(0, start_day_val)

        # Restore edited assignments (convert string keys back to tuples)
        self.edited_assignments = {}
        self.cell_violations = {}  # Violations will be rechecked if needed
        self.cached_timeoff_requests = {}  # Will be repopulated if schedule generates
        for key_str, worker in session.get("edited_assignments", {}).items():
            parts = key_str.split(",")
            if len(parts) == 3:
                try:
                    self.edited_assignments[(int(parts[0]), int(parts[1]), int(parts[2]))] = worker
                except ValueError:
                    pass

        # Validate that all employees in the saved session still exist
        current_employee_names = set(e["name"] for e in self.employees)
        missing_employees = set()

        # Check edited assignments for missing employees
        for cell_key, emp_name in self.edited_assignments.items():
            if emp_name and emp_name not in current_employee_names:
                missing_employees.add(emp_name)

        # Check saved schedule for missing employees
        saved_schedule = session.get("schedule")
        if saved_schedule:
            for key, emp_name in saved_schedule.items():
                if emp_name and emp_name not in current_employee_names:
                    missing_employees.add(emp_name)

        # If there are missing employees, warn the user
        if missing_employees:
            missing_list = ", ".join(sorted(missing_employees))
            response = messagebox.askyesno(
                "Employee List Mismatch",
                f"The saved session includes employees who are no longer in your team:\n\n"
                f"{missing_list}\n\n"
                f"These employees will be removed from the restored schedule.\n\n"
                f"Do you want to continue restoring the session?"
            )
            if not response:
                # User chose not to continue
                self.config_manager.delete_session()
                return

            # Clean up edited_assignments to remove missing employees
            self.edited_assignments = {
                k: v for k, v in self.edited_assignments.items()
                if v in current_employee_names
            }

        if saved_schedule:
            # NEW: Restore schedule directly without solving
            try:
                self.log("Restoring saved schedule…", clear=True)

                month_name = self.month_var.get()
                year = int(self.year_var.get())
                month_num = list(calendar.month_name).index(month_name)

                # Parse holidays and advanced options (same as during generation)
                max_day = calendar.monthrange(year, month_num)[1]
                holidays_list = parse_holidays_string(
                    self.entry_holidays.get().strip(), max_day, self.log
                )
                generate_full_weeks, start_from_day = parse_advanced_options(
                    self.start_day_entry.get().strip(),
                    self.full_weeks_var.get(),
                    max_day, self.log
                )

                # Create generator with same parameters
                gen = ScheduleGeneratorV2(
                    self.employees, year, month_num,
                    skip_weekends=True, holidays=holidays_list,
                    start_from_day=start_from_day,
                    generate_full_weeks=generate_full_weeks
                )

                # Build model to create the shifts dictionary (but don't solve)
                staff_names = [e["name"] for e in self.employees]
                timeoff_raw = {name: entry.get().strip() for name, entry in self.timeoff_entries.items()}
                to_requests = parse_timeoff_entries(
                    timeoff_raw, gen.dates, staff_names, max_day, self.log
                )
                gen.build_model(time_off_requests=to_requests)

                # Replace solver with mock that returns saved values
                gen.solver = MockSolver(saved_schedule, gen.shifts)
                gen.status = cp_model.OPTIMAL  # Mark as solved

                # Render the restored schedule
                self.render_preview(gen)
                self.generator = gen

                # Enable buttons
                self.btn_save.configure(state="normal")
                self.btn_ics.configure(state="normal")
                self.btn_email.configure(state="normal")
                self.btn_save_session.configure(state="normal")

                # Delete session file now that we've restored
                self.config_manager.delete_session()
                # Success! Schedule is now displayed

            except Exception as e:
                # Try to log error, but log panel might be destroyed
                try:
                    self.log(f"  Restore failed: {str(e)}")
                except Exception:
                    pass
                messagebox.showerror(
                    "Restore Failed",
                    f"Failed to restore session:\n{str(e)}\n\n"
                    "You can try generating a new schedule manually."
                )
                # Clear any partial state
                self.edited_assignments = {}
                self.cell_violations = {}
                self.cached_timeoff_requests = {}

        else:
            # OLD: No saved schedule data - re-run solver (backward compatibility)
            try:
                self.log("Re-generating schedule from saved inputs…", clear=True)
                self.run_scheduler()

                # Only delete session if solve succeeded and generator exists
                if self.generator is not None:
                    self.config_manager.delete_session()
                else:
                    # Solve failed - keep session for retry
                    try:
                        self.log("  Session file kept for retry.")
                    except Exception:
                        pass

            except Exception as e:
                # Try to log error, but log panel might be destroyed
                try:
                    self.log(f"  Resume failed: {str(e)}")
                except Exception:
                    pass
                messagebox.showerror(
                    "Resume Failed",
                    f"Failed to restore session:\n{str(e)}\n\n"
                    "You can try generating a new schedule manually."
                )
                # Clear any partial state
                self.edited_assignments = {}
                self.cell_violations = {}
                self.cached_timeoff_requests = {}

    def _show_empty_state(self):
        for widget in self.preview_frame.winfo_children():
            widget.destroy()

        container = ctk.CTkFrame(self.preview_frame, fg_color="transparent")
        container.pack(expand=True, pady=80, padx=40)

        ctk.CTkLabel(
            container,
            text="No schedule generated yet",
            font=make_font(FS_XL, "bold"),
            text_color=TEXT_PRIMARY
        ).pack()
        ctk.CTkLabel(
            container,
            text="Configure options in the sidebar and click Generate Schedule.",
            font=make_font(FS_SM),
            text_color=TEXT_MUTED
        ).pack(pady=(6, 0))

    # ── Log / status messages ─────────────────────────────────────────────────
    def log(self, message, clear=False):
        if clear:
            for widget in self.preview_frame.winfo_children():
                widget.destroy()
            # Progress header
            self._log_header = ctk.CTkFrame(
                self.preview_frame,
                fg_color=SURFACE,
                corner_radius=RADIUS_MD,
                border_width=1, border_color=BORDER
            )
            self._log_header.pack(fill="x", padx=20, pady=(20, 4))

            ctk.CTkLabel(
                self._log_header,
                text="Generating…",
                font=make_font(FS_LG, "bold"),
                text_color=TEXT_PRIMARY
            ).pack(anchor="w", padx=14, pady=(10, 2))

            self._log_body = ctk.CTkFrame(
                self.preview_frame,
                fg_color=SURFACE,
                corner_radius=RADIUS_MD,
                border_width=1, border_color=BORDER
            )
            self._log_body.pack(fill="x", padx=20, pady=(0, 4))

            if hasattr(self, 'preview_canvas'):
                self.preview_canvas.xview_moveto(0)
                self.preview_canvas.yview_moveto(0)

        target = getattr(self, '_log_body', self.preview_frame)
        ctk.CTkLabel(
            target, text=message,
            font=ctk.CTkFont(family="Courier New", size=FS_SM),
            text_color=TEXT_MUTED,
            anchor="w"
        ).pack(fill="x", padx=14, pady=1)
        self.update_idletasks()

    # ── Schedule Preview ──────────────────────────────────────────────────────
    def render_preview(self, generator):
        # Clear cell widget references at the very start to prevent stale references
        self.cell_widgets = {}

        # Save scroll position before destroying children
        x_pos = self.preview_canvas.xview()
        y_pos = self.preview_canvas.yview()

        for widget in self.preview_frame.winfo_children():
            widget.destroy()

        weeks = generator.get_weekly_matrix()
        mode = ctk.get_appearance_mode()
        is_dark = (mode == "Dark")

        def resolve(token):
            if isinstance(token, tuple):
                return token[1] if is_dark else token[0]
            return token

        # Resolved colour vars
        r_header   = ACCENT
        r_phone    = resolve(CELL_PHONE)
        r_fd       = resolve(CELL_FD)
        r_error    = resolve(CELL_ERROR)
        r_empty    = resolve(CELL_EMPTY)
        r_text     = resolve(TEXT_PRIMARY)
        r_err_text = resolve(RED_TEXT)
        r_surface  = resolve(SURFACE)
        r_border   = resolve(BORDER)
        r_muted    = resolve(TEXT_MUTED)

        outer_pad = 20

        # Build double-duty map: {(week_idx, day_idx): {name: True}}
        double_duty_map = {}
        for week_idx, week_data in enumerate(weeks):
            for day_idx in range(5):
                if not week_data["dates"][day_idx]:
                    continue
                # Find who works both phone and FD this day
                workers_phone = set()
                workers_fd = set()
                for s_idx in PHONE_SHIFTS:
                    worker = week_data["matrix"][s_idx][day_idx]
                    if worker:
                        workers_phone.add(worker)
                for s_idx in FD_SHIFTS:
                    worker = week_data["matrix"][s_idx][day_idx]
                    if worker:
                        workers_fd.add(worker)
                # Intersection = people working both
                double_duty_workers = workers_phone & workers_fd
                if double_duty_workers:
                    double_duty_map[(week_idx, day_idx)] = double_duty_workers

        for week_idx, week_data in enumerate(weeks):
            # Week card
            week_card = ctk.CTkFrame(
                self.preview_frame,
                fg_color=r_surface,
                corner_radius=RADIUS_MD,
                border_width=1,
                border_color=r_border
            )
            week_card.pack(fill="x", padx=outer_pad,
                           pady=(outer_pad if week_idx == 0 else 10, 0))

            # Week label row
            week_header = ctk.CTkFrame(
                week_card, fg_color="transparent", corner_radius=0, height=36
            )
            week_header.pack(fill="x", padx=12, pady=(8, 0))

            ctk.CTkLabel(
                week_header,
                text=f"Week {week_idx + 1}",
                font=make_font(FS_LG, "bold"),
                text_color=TEXT_PRIMARY
            ).pack(side="left")

            # Date range hint
            non_empty_dates = [d for d in week_data["dates"] if d]
            if non_empty_dates:
                ctk.CTkLabel(
                    week_header,
                    text=f"  {non_empty_dates[0]} – {non_empty_dates[-1]}",
                    font=make_font(FS_XS),
                    text_color=TEXT_MUTED
                ).pack(side="left", pady=(2, 0))

            # Table grid
            table = ctk.CTkFrame(week_card, fg_color="transparent")
            table.pack(fill="x", padx=10, pady=(6, 12))

            days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

            # ── Column header ──────────────────────────────────────────
            ctk.CTkLabel(
                table,
                text="Shift",
                width=168, height=38,
                fg_color=r_header,
                text_color=TEXT_ON_ACCENT,
                font=make_font(FS_SM, "bold"),
                corner_radius=RADIUS_SM,
                anchor="center"
            ).grid(row=0, column=0, padx=2, pady=2, sticky="nsew")

            for col_i, (day, date_str) in enumerate(zip(days, week_data["dates"])):
                label_txt = f"{day[:3]}\n{date_str}" if date_str else day[:3]
                ctk.CTkLabel(
                    table,
                    text=label_txt,
                    width=128, height=38,
                    fg_color=r_header,
                    text_color=TEXT_ON_ACCENT,
                    font=make_font(FS_SM, "bold"),
                    corner_radius=RADIUS_SM,
                    anchor="center"
                ).grid(row=0, column=col_i + 1, padx=2, pady=2, sticky="nsew")

            # ── Shift rows ─────────────────────────────────────────────
            for r_i, s_idx in enumerate(ALL_SHIFTS):
                shift_name = SHIFT_NAMES[s_idx]

                # Shift label cell
                ctk.CTkLabel(
                    table,
                    text=shift_name,
                    width=168, height=34,
                    fg_color=resolve(SURFACE_ALT),
                    text_color=r_text,
                    font=make_font(FS_SM, "bold"),
                    corner_radius=RADIUS_SM,
                    anchor="w",
                    padx=10
                ).grid(row=r_i + 1, column=0, padx=2, pady=2, sticky="nsew")

                workers = week_data["matrix"][s_idx]

                for c_i, worker in enumerate(workers):
                    display = worker

                    # Check if this cell was manually edited
                    is_manually_edited = (week_idx, c_i, s_idx) in self.edited_assignments
                    if is_manually_edited:
                        display = self.edited_assignments[(week_idx, c_i, s_idx)]

                    has_date = bool(week_data["dates"][c_i])

                    # Check if this day is a holiday
                    is_holiday = week_data.get("holidays", [False] * 5)[c_i]

                    # Check if this person has double duty this day
                    is_double_duty = False
                    if has_date and display:
                        double_duty_workers = double_duty_map.get((week_idx, c_i), set())
                        is_double_duty = display in double_duty_workers

                    if not has_date:
                        bg_c = r_empty
                        txt_c = r_muted
                        cell_txt = ""
                        font_weight = "normal"
                    elif is_holiday:
                        # Holiday cells - light gray background with muted text
                        bg_c = r_empty
                        txt_c = r_muted
                        cell_txt = "Holiday"
                        font_weight = "normal"
                    elif not display:
                        bg_c = r_error
                        txt_c = r_err_text
                        cell_txt = "UNASSIGNED"
                        font_weight = "normal"
                    elif s_idx in PHONE_SHIFTS:
                        bg_c = r_phone
                        txt_c = r_text
                        cell_txt = f"{display} !" if is_double_duty else display
                        font_weight = "bold" if is_double_duty else "normal"
                    else:  # FD
                        bg_c = r_fd
                        txt_c = r_text
                        cell_txt = f"{display} !" if is_double_duty else display
                        font_weight = "bold" if is_double_duty else "normal"

                    # Check for violations
                    cell_key = (week_idx, c_i, s_idx)
                    has_violation = cell_key in self.cell_violations

                    # Always create cells with a border frame for flexibility
                    # Determine border color
                    if has_violation:
                        border_color = "#EF4444"  # red-500 for violations
                        border_width = 2
                    elif is_manually_edited:
                        border_color = ACCENT  # blue for manual edits
                        border_width = 2
                    else:
                        border_color = "transparent"
                        border_width = 0

                    # Create border frame
                    border_frame = ctk.CTkFrame(
                        table,
                        fg_color=border_color if border_width > 0 else "transparent",
                        corner_radius=RADIUS_SM,
                        border_width=0
                    )
                    border_frame.grid(row=r_i + 1, column=c_i + 1,
                                     padx=2, pady=2, sticky="nsew")

                    # Create inner label
                    lbl = ctk.CTkLabel(
                        border_frame,
                        text=cell_txt,
                        fg_color=bg_c,
                        text_color=txt_c,
                        font=make_font(FS_SM, font_weight),
                        corner_radius=RADIUS_SM - 1 if border_width > 0 else RADIUS_SM,
                        anchor="center",
                        width=128 if border_width == 0 else None,
                        height=34 if border_width == 0 else None
                    )
                    if border_width > 0:
                        lbl.pack(fill="both", expand=True, padx=3, pady=3)
                    else:
                        lbl.pack(fill="both", expand=True, padx=2, pady=2)

                    # Store both label and border frame for incremental updates
                    self.cell_widgets[(week_idx, c_i, s_idx)] = (lbl, border_frame)

                    # Make cells clickable only if they have a date and are not holidays
                    if has_date and not is_holiday:
                        lbl.bind(
                            "<Button-1>",
                            lambda e, w=week_idx, d=c_i, s=s_idx, n=display:
                                self.on_cell_click(e, w, d, s, n)
                        )

                        # Add violation tooltip or cursor change based on state
                        if has_violation:
                            viol_type, conflict_key, msg = self.cell_violations[cell_key]

                            # Create tooltip functions with cell_key in closure
                            def make_show_tooltip(lbl_ref, message):
                                def show_tooltip(event):
                                    lbl_ref.configure(cursor="hand2")
                                    # Create tooltip
                                    tooltip = ctk.CTkFrame(self, fg_color="#1F2937", corner_radius=4,
                                                          border_width=1, border_color="#374151")
                                    tooltip_label = ctk.CTkLabel(
                                        tooltip, text=message,
                                        font=make_font(FS_XS, "normal"),
                                        text_color="#F9FAFB",
                                        padx=8, pady=4
                                    )
                                    tooltip_label.pack()

                                    # Position tooltip near cursor
                                    x = event.x_root + 10
                                    y = event.y_root + 10
                                    tooltip.place(x=x - self.winfo_rootx(), y=y - self.winfo_rooty())

                                    # Store tooltip reference
                                    lbl_ref.tooltip = tooltip
                                return show_tooltip

                            def make_hide_tooltip(lbl_ref):
                                def hide_tooltip(event):
                                    lbl_ref.configure(cursor="")
                                    if hasattr(lbl_ref, 'tooltip'):
                                        lbl_ref.tooltip.destroy()
                                        delattr(lbl_ref, 'tooltip')
                                return hide_tooltip

                            lbl.bind("<Enter>", make_show_tooltip(lbl, msg))
                            lbl.bind("<Leave>", make_hide_tooltip(lbl))
                        else:
                            lbl.bind("<Enter>", lambda e, l=lbl: l.configure(cursor="hand2"))
                            lbl.bind("<Leave>", lambda e, l=lbl: l.configure(cursor=""))

                            # Add tooltip for double duty
                            if is_double_duty:
                                ToolTip(lbl, f"{display} has TWO shifts this day\n(Phone + Front Desk)")

        # Render fairness audit in the side panel
        self._render_fairness_panel(generator)

        # Rebind scroll events to all newly created widgets
        self._bind_canvas_scroll()

        # Restore scroll position after rebuild
        def _restore_scroll():
            self.preview_canvas.xview_moveto(x_pos[0])
            self.preview_canvas.yview_moveto(y_pos[0])
        self.after(50, _restore_scroll)

    def update_single_cell(self, week_idx, day_idx, shift_idx):
        """Update a single cell without re-rendering the entire schedule."""
        # Check if we have a widget reference for this cell
        cell_key = (week_idx, day_idx, shift_idx)
        if cell_key not in self.cell_widgets:
            # Widget doesn't exist yet, do nothing (fallback to full render if needed)
            return

        lbl, border_frame = self.cell_widgets[cell_key]

        # Get current data from generator
        weeks = self.generator.get_weekly_matrix()
        if week_idx >= len(weeks):
            return

        week_data = weeks[week_idx]

        # Bounds checking for shift_idx and day_idx
        if shift_idx >= len(week_data["matrix"]) or shift_idx < 0:
            return
        if day_idx >= len(week_data["matrix"][shift_idx]) or day_idx < 0:
            return

        worker = week_data["matrix"][shift_idx][day_idx]

        # Resolve appearance mode colors
        mode = ctk.get_appearance_mode()
        is_dark = (mode == "Dark")

        def resolve(token):
            if isinstance(token, tuple):
                return token[1] if is_dark else token[0]
            return token

        r_phone = resolve(CELL_PHONE)
        r_fd = resolve(CELL_FD)
        r_error = resolve(CELL_ERROR)
        r_empty = resolve(CELL_EMPTY)
        r_text = resolve(TEXT_PRIMARY)
        r_err_text = resolve(RED_TEXT)
        r_muted = resolve(TEXT_MUTED)

        # Check if manually edited
        is_manually_edited = cell_key in self.edited_assignments
        display = self.edited_assignments[cell_key] if is_manually_edited else worker

        # Bounds check for dates array
        if day_idx >= len(week_data["dates"]):
            return
        has_date = bool(week_data["dates"][day_idx])

        # Check if this day is a holiday (with bounds check)
        holidays = week_data.get("holidays", [False] * 5)
        if day_idx >= len(holidays):
            return
        is_holiday = holidays[day_idx]

        # Check double duty
        is_double_duty = False
        if has_date and display:
            # Build double duty map for this specific day
            workers_phone = set()
            workers_fd = set()
            for s_idx in PHONE_SHIFTS:
                w = week_data["matrix"][s_idx][day_idx]
                # Check if this cell was manually edited
                if (week_idx, day_idx, s_idx) in self.edited_assignments:
                    w = self.edited_assignments[(week_idx, day_idx, s_idx)]
                if w:
                    workers_phone.add(w)
            for s_idx in FD_SHIFTS:
                w = week_data["matrix"][s_idx][day_idx]
                # Check if this cell was manually edited
                if (week_idx, day_idx, s_idx) in self.edited_assignments:
                    w = self.edited_assignments[(week_idx, day_idx, s_idx)]
                if w:
                    workers_fd.add(w)
            double_duty_workers = workers_phone & workers_fd
            is_double_duty = display in double_duty_workers

        # Determine cell appearance
        if not has_date:
            bg_c = r_empty
            txt_c = r_muted
            cell_txt = ""
            font_weight = "normal"
        elif is_holiday:
            # Holiday cells - light gray background with muted text
            bg_c = r_empty
            txt_c = r_muted
            cell_txt = "Holiday"
            font_weight = "normal"
        elif not display:
            bg_c = r_error
            txt_c = r_err_text
            cell_txt = "UNASSIGNED"
            font_weight = "normal"
        elif shift_idx in PHONE_SHIFTS:
            bg_c = r_phone
            txt_c = r_text
            cell_txt = f"{display} !" if is_double_duty else display
            font_weight = "bold" if is_double_duty else "normal"
        else:  # FD
            bg_c = r_fd
            txt_c = r_text
            cell_txt = f"{display} !" if is_double_duty else display
            font_weight = "bold" if is_double_duty else "normal"

        # Check widget still exists before updating
        if not lbl.winfo_exists() or not border_frame.winfo_exists():
            return

        # Check for violations
        has_violation = cell_key in self.cell_violations
        if has_violation:
            viol_type, conflict_key, msg = self.cell_violations[cell_key]
            # Set red border for violations
            border_frame.configure(fg_color="#EF4444")  # red-500

            # Add tooltip on hover
            def show_tooltip(event):
                if not lbl.winfo_exists():
                    return
                lbl.configure(cursor="hand2")
                # Create tooltip
                tooltip = ctk.CTkFrame(self, fg_color="#1F2937", corner_radius=4,
                                      border_width=1, border_color="#374151")
                tooltip_label = ctk.CTkLabel(
                    tooltip, text=msg,
                    font=make_font(FS_XS, "normal"),
                    text_color="#F9FAFB",
                    padx=8, pady=4
                )
                tooltip_label.pack()

                # Position tooltip near cursor
                x = event.x_root + 10
                y = event.y_root + 10
                tooltip.place(x=x - self.winfo_rootx(), y=y - self.winfo_rooty())

                # Store tooltip reference to remove it later
                lbl.tooltip = tooltip

            def hide_tooltip(event):
                if not lbl.winfo_exists():
                    return
                lbl.configure(cursor="")
                if hasattr(lbl, 'tooltip') and lbl.tooltip.winfo_exists():
                    lbl.tooltip.destroy()
                    delattr(lbl, 'tooltip')

            # Bind hover events
            lbl.bind("<Enter>", show_tooltip)
            lbl.bind("<Leave>", hide_tooltip)
        else:
            # Reset border to normal (blue for manual edits, or transparent)
            if is_manually_edited:
                border_frame.configure(fg_color=ACCENT)
            else:
                border_frame.configure(fg_color="transparent")

            # Restore normal cursor bindings
            lbl.bind("<Enter>", lambda e: lbl.configure(cursor="hand2") if lbl.winfo_exists() else None)
            lbl.bind("<Leave>", lambda e: lbl.configure(cursor="") if lbl.winfo_exists() else None)

            # Clean up tooltip if it exists
            if hasattr(lbl, 'tooltip') and lbl.tooltip.winfo_exists():
                lbl.tooltip.destroy()
                delattr(lbl, 'tooltip')

        # Update the label
        lbl.configure(
            text=cell_txt,
            fg_color=bg_c,
            text_color=txt_c,
            font=make_font(FS_SM, font_weight)
        )

    # ── Cell Edit (inline panel) ─────────────────────────────────────────────
    def on_cell_click(self, event, week_idx, day_idx, shift_idx, current_name):
        def build(content):
            staff_names = [e["name"] for e in self.employees]
            options = ["UNASSIGNED"] + sorted(staff_names)  # Removed "" duplicate

            ctk.CTkLabel(
                content, text=SHIFT_NAMES[shift_idx],
                font=make_font(FS_BASE, "bold"),
                text_color=TEXT_PRIMARY
            ).pack(anchor="w", pady=(0, 2))

            ctk.CTkLabel(
                content, text="Enter a name or pick from the roster:",
                font=make_font(FS_XS), text_color=TEXT_MUTED
            ).pack(anchor="w", pady=(0, 6))

            selection = ctk.StringVar(value=current_name if current_name else "UNASSIGNED")

            custom_entry = ctk.CTkEntry(
                content, placeholder_text="Custom name…",
                font=make_font(FS_SM),
                height=34, corner_radius=RADIUS_SM,
                border_color=BORDER
            )
            custom_entry.pack(fill="x", pady=(0, 6))
            if current_name and current_name not in options:
                custom_entry.insert(0, current_name)

            def save_and_close():
                custom_val = custom_entry.get().strip()
                val = custom_val if custom_val else selection.get()
                if val == "UNASSIGNED":
                    val = ""
                self.edited_assignments[(week_idx, day_idx, shift_idx)] = val
                self._hide_action_panel()

                # Collect all cells that need to be rechecked
                cells_to_recheck = set()

                # Add the edited cell
                cells_to_recheck.add((week_idx, day_idx, shift_idx))

                # Add all cells on the same day (for double duty / overlap checks)
                for s_idx in ALL_SHIFTS:
                    cell_key = (week_idx, day_idx, s_idx)
                    cells_to_recheck.add(cell_key)

                # Add any cells that were previously marked as conflicting with this cell
                # (They may no longer have violations after this edit)
                cells_with_violations_to_clear = []
                for viol_cell_key, (viol_type, conflict_key, msg) in list(self.cell_violations.items()):
                    if conflict_key == (week_idx, day_idx, shift_idx):
                        cells_to_recheck.add(viol_cell_key)
                        cells_with_violations_to_clear.append(viol_cell_key)

                # Clear violations for all cells we're about to recheck
                for cell_key in cells_to_recheck:
                    if cell_key in self.cell_violations:
                        del self.cell_violations[cell_key]

                # Re-check all affected cells for violations
                for check_cell in cells_to_recheck:
                    check_w, check_d, check_s = check_cell
                    if check_cell in self.edited_assignments:
                        check_name = self.edited_assignments[check_cell]
                        violations = self.check_manual_edit_violations(check_w, check_d, check_s, check_name)

                        # Store new violations
                        if violations:
                            for viol_type, conflict_key, msg in violations:
                                self.cell_violations[check_cell] = (viol_type, conflict_key, msg)
                                # Also mark the conflicting cell (whether manually edited or not)
                                if conflict_key:
                                    self.cell_violations[conflict_key] = (viol_type, check_cell, msg)
                                    # Add conflicting cell to recheck list to update its visual
                                    if conflict_key not in cells_to_recheck:
                                        cells_to_recheck.add(conflict_key)
                                break  # Only store first violation per cell

                # Update only the affected cells instead of re-rendering everything
                for check_cell in cells_to_recheck:
                    self.update_single_cell(check_cell[0], check_cell[1], check_cell[2])

                # Also update the fairness panel
                self._render_fairness_panel(self.generator)

            ctk.CTkButton(
                content, text="Save Change",
                command=save_and_close,
                fg_color=ACCENT, hover_color=ACCENT_HOVER,
                text_color=TEXT_ON_ACCENT,
                font=make_font(FS_SM, "bold"),
                height=34, corner_radius=RADIUS_SM
            ).pack(fill="x", pady=(0, 8))

            ctk.CTkFrame(content, height=1, fg_color=BORDER,
                         corner_radius=0).pack(fill="x", pady=(0, 6))

            ctk.CTkLabel(
                content, text="Or select from roster",
                font=make_font(FS_XS), text_color=TEXT_MUTED
            ).pack(anchor="w", pady=(0, 4))

            for opt in options:
                ctk.CTkRadioButton(
                    content, text=opt,
                    variable=selection, value=opt,
                    font=make_font(FS_SM),
                    fg_color=ACCENT,
                    command=save_and_close
                ).pack(anchor="w", pady=2)

        self._show_action_panel("Edit Shift Assignment", build)

    # ── Fairness Info (inline panel) ─────────────────────────────────────────
    def show_settings_menu(self):
        """Display a settings menu with all configuration options."""
        def build(content):
            # Manage Team option
            team_card = ctk.CTkFrame(
                content, fg_color=SURFACE_ALT,
                corner_radius=RADIUS_SM,
                border_width=1, border_color=BORDER
            )
            team_card.pack(fill="x", pady=(0, 8))

            ctk.CTkLabel(
                team_card, text="Manage Team",
                font=make_font(FS_SM, "bold"),
                text_color=TEXT_PRIMARY, anchor="w"
            ).pack(anchor="w", padx=12, pady=(10, 4))

            ctk.CTkLabel(
                team_card, text="Add, edit, or remove employees from your roster.",
                font=make_font(FS_XS),
                text_color=TEXT_MUTED,
                anchor="w", wraplength=300, justify="left"
            ).pack(anchor="w", padx=12, pady=(0, 8))

            ctk.CTkButton(
                team_card, text="Open Team Manager",
                command=self.open_team_manager,
                fg_color=SURFACE, hover_color=SURFACE_ALT,
                text_color=TEXT_PRIMARY,
                border_width=1, border_color=BORDER,
                font=make_font(FS_XS, "bold"),
                height=28, corner_radius=RADIUS_SM
            ).pack(fill="x", padx=12, pady=(0, 10))

            # Fairness Info option
            fairness_card = ctk.CTkFrame(
                content, fg_color=SURFACE_ALT,
                corner_radius=RADIUS_SM,
                border_width=1, border_color=BORDER
            )
            fairness_card.pack(fill="x", pady=(0, 8))

            ctk.CTkLabel(
                fairness_card, text="Fairness Info",
                font=make_font(FS_SM, "bold"),
                text_color=TEXT_PRIMARY, anchor="w"
            ).pack(anchor="w", padx=12, pady=(10, 4))

            ctk.CTkLabel(
                fairness_card, text="Learn how the schedule generator ensures fair shift distribution.",
                font=make_font(FS_XS),
                text_color=TEXT_MUTED,
                anchor="w", wraplength=300, justify="left"
            ).pack(anchor="w", padx=12, pady=(0, 8))

            ctk.CTkButton(
                fairness_card, text="View Fairness Rules",
                command=self.show_about_dialog,
                fg_color=SURFACE, hover_color=SURFACE_ALT,
                text_color=TEXT_PRIMARY,
                border_width=1, border_color=BORDER,
                font=make_font(FS_XS, "bold"),
                height=28, corner_radius=RADIUS_SM
            ).pack(fill="x", padx=12, pady=(0, 10))

            # Reset Email Settings option
            email_card = ctk.CTkFrame(
                content, fg_color=SURFACE_ALT,
                corner_radius=RADIUS_SM,
                border_width=1, border_color=BORDER
            )
            email_card.pack(fill="x", pady=(0, 0))

            ctk.CTkLabel(
                email_card, text="Reset Email Settings",
                font=make_font(FS_SM, "bold"),
                text_color=TEXT_PRIMARY, anchor="w"
            ).pack(anchor="w", padx=12, pady=(10, 4))

            ctk.CTkLabel(
                email_card, text="Clear saved Outlook settings and re-run verification on next email send.",
                font=make_font(FS_XS),
                text_color=TEXT_MUTED,
                anchor="w", wraplength=300, justify="left"
            ).pack(anchor="w", padx=12, pady=(0, 8))

            ctk.CTkButton(
                email_card, text="Reset Email Settings",
                command=lambda: [self._hide_action_panel(), self._reset_email_settings()],
                fg_color=SURFACE, hover_color=SURFACE_ALT,
                text_color=TEXT_PRIMARY,
                border_width=1, border_color=BORDER,
                font=make_font(FS_XS, "bold"),
                height=28, corner_radius=RADIUS_SM
            ).pack(fill="x", padx=12, pady=(0, 10))

            # Developer Credit
            ctk.CTkLabel(
                content, text="Developed by CyberNabbed",
                font=make_font(FS_XS),
                text_color=TEXT_MUTED,
                anchor="center"
            ).pack(fill="x", pady=(8, 0))

        self._show_action_panel("Settings", build)

    def show_about_dialog(self):
        def build(content):
            points = [
                ("REQUIRED RULES - Cannot Be Broken",
                 "These rules are absolute. If violated, the schedule will fail to generate."),
                ("No Overlapping Shifts",
                 "You cannot work two shifts that overlap in time on the same day. The engine blocks these combinations automatically."),
                ("Availability Matches Schedule Type",
                 "Late starters cannot be assigned the 7:30am shift. Early leavers cannot be assigned the shift ending at 5pm."),
                ("Daily Workload Limits",
                 "Maximum 1 phone shift per day. Maximum 1 front desk shift per day. No exceptions."),
                ("Every Shift Gets Covered",
                 "Each shift must have exactly one person assigned. Holidays are the only exception where shifts are left empty."),
                ("Fair Distribution Based on Availability",
                 "Everyone works approximately the same number of shifts, adjusted for their availability. If you take more time off, you work proportionally fewer shifts. You'll always be within 1 shift of your fair share."),
                ("PREFERENCES - The Engine Tries to Honor These",
                 "These improve schedule quality but may be compromised if necessary."),
                ("Strongest: Avoid Double Duty",
                 "The engine really dislikes assigning phone + front desk on the same day. This gets the highest penalty weight."),
                ("Extra Strong: Avoid Back-to-Back Double Duty",
                 "If you must work both phone and front desk, the engine strongly prefers a break between them. P2→FD2 and FD1→P3 (no break) get extra heavy penalties."),
                ("Strong: Space Out Front Desk Days",
                 "The engine tries hard to avoid scheduling front desk shifts on back-to-back days."),
                ("Strong: Share Double Duty Fairly",
                 "If double duty is unavoidable, the engine distributes it evenly across the team rather than overloading specific people."),
                ("Lower: Variety in Shifts",
                 "The engine prefers not to give you the same shift two days in a row, but this is the lowest priority preference."),
                ("HOW IT WORKS",
                 "The engine satisfies all required rules first, then optimizes to honor as many preferences as possible in priority order."),
            ]

            for i, (title, detail) in enumerate(points):
                # Determine styling based on position
                is_header = i in [0, 6, 12]  # Headers
                is_required = 1 <= i <= 5    # Required rules
                is_preference = 7 <= i <= 11 # Preferences

                if is_header:
                    # Header styling
                    card = ctk.CTkFrame(
                        content, fg_color=ACCENT_LIGHT,
                        corner_radius=RADIUS_SM,
                        border_width=0
                    )
                    card.pack(fill="x", pady=(12 if i > 0 else 0, 4))

                    ctk.CTkLabel(
                        card, text=title,
                        font=make_font(FS_SM, "bold"),
                        text_color=TEXT_PRIMARY, anchor="w"
                    ).pack(anchor="w", padx=10, pady=8)

                    ctk.CTkLabel(
                        card, text=detail,
                        font=make_font(FS_XS),
                        text_color=TEXT_MUTED,
                        anchor="w", wraplength=300, justify="left"
                    ).pack(anchor="w", padx=10, pady=(0, 8))

                elif is_required:
                    # Required rules - red/orange tint
                    card = ctk.CTkFrame(
                        content, fg_color=SURFACE_ALT,
                        corner_radius=RADIUS_SM,
                        border_width=2, border_color=("#EF4444", "#7F1D1D")  # red
                    )
                    card.pack(fill="x", pady=(0, 4))

                    ctk.CTkLabel(
                        card, text=title,
                        font=make_font(FS_XS, "bold"),
                        text_color=TEXT_PRIMARY, anchor="w"
                    ).pack(anchor="w", padx=8, pady=(6, 0))

                    ctk.CTkLabel(
                        card, text=detail,
                        font=make_font(FS_XS),
                        text_color=TEXT_MUTED,
                        anchor="w", wraplength=300, justify="left"
                    ).pack(anchor="w", padx=8, pady=(2, 6))

                elif is_preference:
                    # Preferences - blue/green tint
                    card = ctk.CTkFrame(
                        content, fg_color=SURFACE_ALT,
                        corner_radius=RADIUS_SM,
                        border_width=2, border_color=("#10B981", "#065F46")  # green
                    )
                    card.pack(fill="x", pady=(0, 4))

                    ctk.CTkLabel(
                        card, text=title,
                        font=make_font(FS_XS, "bold"),
                        text_color=TEXT_PRIMARY, anchor="w"
                    ).pack(anchor="w", padx=8, pady=(6, 0))

                    ctk.CTkLabel(
                        card, text=detail,
                        font=make_font(FS_XS),
                        text_color=TEXT_MUTED,
                        anchor="w", wraplength=300, justify="left"
                    ).pack(anchor="w", padx=8, pady=(2, 6))

        self._show_action_panel("How Fairness Works", build, back_command=self.show_settings_menu)

    # ── Team Manager (inline panel) ─────────────────────────────────────────
    def open_team_manager(self):
        def build(content):
            section_label(content, "Current Roster").pack(anchor="w", pady=(0, 4))

            roster_frame = ctk.CTkFrame(content, fg_color="transparent")
            roster_frame.pack(fill="x")

            def refresh_roster():
                for w in roster_frame.winfo_children():
                    w.destroy()
                for emp in self.employees:
                    self._render_roster_row(roster_frame, emp, None, refresh_roster)

            refresh_roster()

            ctk.CTkFrame(content, height=1, fg_color=BORDER,
                         corner_radius=0).pack(fill="x", pady=(8, 6))

            section_label(content, "Add New Agent").pack(anchor="w", pady=(0, 4))

            e_name = ctk.CTkEntry(
                content, placeholder_text="Full name…",
                font=make_font(FS_SM),
                height=32, corner_radius=RADIUS_SM, border_color=BORDER
            )
            e_name.pack(fill="x", pady=(0, 4))

            e_email = ctk.CTkEntry(
                content, placeholder_text="Email address…",
                font=make_font(FS_SM),
                height=32, corner_radius=RADIUS_SM, border_color=BORDER
            )
            e_email.pack(fill="x", pady=(0, 4))

            e_type = ctk.StringVar(value="early")
            type_row = ctk.CTkFrame(content, fg_color="transparent")
            type_row.pack(fill="x", pady=(0, 6))
            ctk.CTkRadioButton(type_row, text="Early",
                               variable=e_type, value="early",
                               font=make_font(FS_XS)).pack(side="left", padx=(0, 10))
            ctk.CTkRadioButton(type_row, text="Late",
                               variable=e_type, value="late",
                               font=make_font(FS_XS)).pack(side="left")

            def do_add():
                self.add_emp(e_name.get(), e_type.get(), None, e_email.get())
                e_name.delete(0, "end")
                e_email.delete(0, "end")
                refresh_roster()

            e_email.bind("<Return>", lambda ev: do_add())

            ctk.CTkButton(
                content, text="Add Employee",
                command=do_add,
                fg_color=ACCENT, hover_color=ACCENT_HOVER,
                text_color=TEXT_ON_ACCENT,
                font=make_font(FS_SM, "bold"),
                height=34, corner_radius=RADIUS_SM
            ).pack(fill="x", pady=(0, 4))

        self._show_action_panel("Manage Team", build, back_command=self.show_settings_menu)

    def _render_roster_row(self, parent, emp, mgr_window, refresh_fn):
        name = emp["name"]
        etype = emp.get("type", "early")
        email = emp.get("email", "")

        row = ctk.CTkFrame(
            parent, fg_color=SURFACE,
            corner_radius=RADIUS_SM,
            border_width=1, border_color=BORDER
        )
        row.pack(fill="x", pady=3)
        row.grid_columnconfigure(1, weight=1)

        badge_color = BLUE_BG if etype == "early" else ORANGE_BG
        badge_text_color = BLUE_TEXT if etype == "early" else ORANGE_TEXT

        ctk.CTkLabel(
            row, text=name[0].upper(),
            width=30, height=30,
            fg_color=badge_color, text_color=badge_text_color,
            corner_radius=15, font=make_font(FS_XS, "bold")
        ).grid(row=0, column=0, rowspan=2, padx=(10, 8), pady=8)

        ctk.CTkLabel(
            row, text=name,
            font=make_font(FS_SM, "bold"),
            text_color=TEXT_PRIMARY, anchor="w"
        ).grid(row=0, column=1, sticky="sw")

        email_display = email if email else "(no email)"
        ctk.CTkLabel(
            row, text=email_display,
            font=make_font(FS_XS),
            text_color=TEXT_MUTED, anchor="w"
        ).grid(row=1, column=1, sticky="nw")

        ctk.CTkLabel(
            row, text=etype.capitalize(),
            font=make_font(FS_XS),
            text_color=TEXT_MUTED, anchor="w"
        ).grid(row=0, column=2, rowspan=2, padx=(0, 8))

        def remove():
            if messagebox.askyesno(
                "Remove Employee",
                f"Remove {name} from the team?\nThis cannot be undone."
            ):
                self.remove_emp(emp, mgr_window, refresh_fn)

        ctk.CTkButton(
            row, text="Remove",
            command=remove,
            fg_color="transparent",
            hover_color=RED_BG,
            text_color=RED_TEXT,
            border_width=1, border_color=RED_TEXT,
            font=make_font(FS_XS),
            width=60, height=26,
            corner_radius=RADIUS_SM
        ).grid(row=0, column=3, rowspan=2, padx=(0, 10))

    def remove_emp(self, emp, window, refresh_fn=None):
        emp_name = emp["name"]
        self.employees.remove(emp)
        self.config_manager.save_config(self.employees)
        self.setup_timeoff_ui()

        # Clean up manual edits and violations for this employee
        cells_to_update = []

        # Remove from edited_assignments
        for cell_key, assigned_name in list(self.edited_assignments.items()):
            if assigned_name == emp_name:
                del self.edited_assignments[cell_key]
                cells_to_update.append(cell_key)

        # Remove from cell_violations
        for cell_key in list(self.cell_violations.keys()):
            # Remove violations where this employee is involved
            if cell_key in self.edited_assignments and self.edited_assignments[cell_key] == emp_name:
                del self.cell_violations[cell_key]
            # Also check conflict keys
            viol_type, conflict_key, msg = self.cell_violations.get(cell_key, (None, None, None))
            if conflict_key and conflict_key in self.edited_assignments and self.edited_assignments[conflict_key] == emp_name:
                del self.cell_violations[cell_key]

        # Update displayed cells if schedule exists
        if self.generator and cells_to_update:
            for cell_key in cells_to_update:
                week_idx, day_idx, shift_idx = cell_key
                self.update_single_cell(week_idx, day_idx, shift_idx)
            # Update fairness panel
            self._render_fairness_panel(self.generator)

        if refresh_fn:
            refresh_fn()
        elif window:
            window.destroy()
            self.open_team_manager()

    def add_emp(self, name, etype, window, email=""):
        name = name.strip()
        email = email.strip() if email else ""
        is_valid, error_msg = self.validate_employee_name(name)
        if not is_valid:
            messagebox.showwarning("Input Error", error_msg)
            return
        if email and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
            messagebox.showwarning("Input Error", "Please enter a valid email address.")
            return
        self.employees.append({"name": name, "type": etype, "email": email})
        try:
            self.config_manager.save_config(self.employees)
        except Exception as e:
            self.employees.pop()
            messagebox.showerror("Save Error", f"Failed to save: {str(e)}")
            return
        self.setup_timeoff_ui()
        if window:
            window.destroy()
            self.open_team_manager()

    # ── Schedule Generation ───────────────────────────────────────────────────
    def run_scheduler(self):
        # Check if a schedule already exists and ask for confirmation
        if self.generator is not None:
            response = messagebox.askyesnocancel(
                "Overwrite Schedule?",
                "A schedule is currently displayed.\n\n"
                "Generating a new schedule will overwrite your current schedule "
                "and any manual edits you've made.\n\n"
                "Do you want to continue?"
            )
            # response: True = Yes, False = No, None = Cancel
            if response is None or response is False:
                # User chose Cancel or No, don't generate
                return

        # Disable generate button to prevent multiple concurrent generations
        self.btn_generate.configure(state="disabled")

        try:
            self.employees = self.config_manager.get_employees()
            if len(self.employees) < 6:
                messagebox.showwarning(
                    "Low Staff",
                    "You have fewer than 6 employees configured.\n"
                    "Coverage requirements (6 shifts/day) may not be met."
                )
            elif len(self.employees) > 50:
                if not messagebox.askyesno(
                    "Large Team Warning",
                    f"You have {len(self.employees)} employees.\n"
                    "This may cause slow generation. Continue?"
                ):
                    self.btn_generate.configure(state="normal")
                    return

            month_name = self.month_var.get()
            year = int(self.year_var.get())
            month_num = list(calendar.month_name).index(month_name)

            self.log(
                f"Initialising  {month_name} {year}  ·  "
                f"{len(self.employees)} employees…",
                clear=True
            )

            # Clear manual edits and violations
            self.edited_assignments = {}
            self.cell_violations = {}
            self.cached_timeoff_requests = {}  # Will be repopulated below

            # Parse holidays and advanced options
            max_day = calendar.monthrange(year, month_num)[1]
            holidays_list = parse_holidays_string(
                self.entry_holidays.get().strip(), max_day, self.log
            )
            if holidays_list:
                self.log(f"  Holidays: {sorted(set(holidays_list))}")

            generate_full_weeks, start_from_day = parse_advanced_options(
                self.start_day_entry.get().strip(),
                self.full_weeks_var.get(),
                max_day, self.log
            )

            gen = ScheduleGeneratorV2(
                self.employees, year, month_num,
                skip_weekends=True, holidays=holidays_list,
                start_from_day=start_from_day,
                generate_full_weeks=generate_full_weeks
            )

            # Process time-off requests FIRST (before building model)
            self.log("  Processing time-off requests…")
            staff_names = [e["name"] for e in self.employees]
            timeoff_raw = {name: entry.get().strip() for name, entry in self.timeoff_entries.items()}
            to_requests = parse_timeoff_entries(
                timeoff_raw, gen.dates, staff_names, max_day, self.log
            )

            # Cache time-off snapshot for violation checking
            self.cached_timeoff_requests = to_requests.copy()

            # Build model with time-off awareness for proportional fairness
            self.log("  Building schedule model…")
            gen.build_model(time_off_requests=to_requests)

            # Apply time-off constraints (force shifts to 0 on time-off days)
            for (name, idx) in to_requests.keys():
                for s in ALL_SHIFTS:
                    gen.model.Add(gen.shifts[(name, idx, s)] == 0)

            self.log("  Solving…")
            success = gen.solve()

            if success:
                self.render_preview(gen)
                self.generator = gen
                self.btn_save.configure(state="normal")
                self.btn_ics.configure(state="normal")
                self.btn_email.configure(state="normal")
                self.btn_save_session.configure(state="normal")
                self.btn_generate.configure(state="normal")
            else:
                self.log("\n  No solution found — diagnosing constraints…")
                issues = gen.diagnose(to_requests)

                error_msg = "No valid schedule could be built.\n\nCritical issues:\n"
                display_limit = 8
                for i, issue in enumerate(issues):
                    self.log(f"  ! {issue}")
                    if i < display_limit:
                        error_msg += f"• {issue}\n"
                if len(issues) > display_limit:
                    error_msg += f"…and {len(issues) - display_limit} more (see log).\n"
                if not issues:
                    error_msg += "• Unknown constraint conflict.\n"

                error_msg += (
                    "\nGenerate a partial schedule anyway?\n"
                    "(Unresolvable slots will be left empty in red.)"
                )

                if messagebox.askyesno("Constraints Too Tight", error_msg):
                    self.log("\n  Retrying with partial coverage…")
                    gen = ScheduleGeneratorV2(
                        self.employees, year, month_num,
                        skip_weekends=True, holidays=holidays_list,
                        start_from_day=start_from_day,
                        generate_full_weeks=generate_full_weeks
                    )
                    gen.build_model(time_off_requests=to_requests, soft_coverage=True)
                    for (name, idx) in to_requests.keys():
                        for s in ALL_SHIFTS:
                            gen.model.Add(gen.shifts[(name, idx, s)] == 0)
                    if gen.solve():
                        self.render_preview(gen)
                        self.generator = gen
                        self.btn_save.configure(state="normal")
                        self.btn_ics.configure(state="normal")
                        self.btn_email.configure(state="normal")
                        self.btn_generate.configure(state="normal")
                    else:
                        self.log("  Failed even with partial coverage.")
                        self.btn_generate.configure(state="normal")
                else:
                    # User declined partial schedule
                    self.btn_generate.configure(state="normal")

        except Exception as e:
            self.log(f"  Error: {str(e)}")
            messagebox.showerror("Generation Error", str(e))
            import traceback
            traceback.print_exc()
            self.btn_generate.configure(state="normal")

    # ── Export Calendar (inline panel) ───────────────────────────────────────
    def export_calendar_dialog(self):
        if not self.generator:
            return

        def build(content):
            ctk.CTkLabel(
                content, text="Choose an export mode:",
                font=make_font(FS_SM, "bold"), text_color=TEXT_PRIMARY
            ).pack(anchor="w", pady=(0, 8))

            mode_var = ctk.StringVar(value="separate")

            ctk.CTkRadioButton(
                content, text="Separate file per agent",
                variable=mode_var, value="separate",
                font=make_font(FS_SM)
            ).pack(anchor="w", pady=(0, 2))
            ctk.CTkLabel(
                content, text="One .ics per team member.",
                font=make_font(FS_XS), text_color=TEXT_MUTED
            ).pack(anchor="w", padx=(20, 0), pady=(0, 8))

            ctk.CTkRadioButton(
                content, text="Full schedule (one file)",
                variable=mode_var, value="full",
                font=make_font(FS_SM)
            ).pack(anchor="w", pady=(0, 2))
            ctk.CTkLabel(
                content, text="Single .ics with all shifts.",
                font=make_font(FS_XS), text_color=TEXT_MUTED
            ).pack(anchor="w", padx=(20, 0), pady=(0, 10))

            def do_export():
                folder = filedialog.askdirectory(title="Choose folder for calendar files")
                if not folder:
                    return
                try:
                    month_name = self.month_var.get()
                    year = int(self.year_var.get())
                    month_num = list(calendar.month_name).index(month_name)
                    created = generate_all_ics(
                        self.generator, self.employees, self.edited_assignments,
                        year, month_num, folder, mode=mode_var.get()
                    )
                    if created:
                        file_list = "\n".join(os.path.basename(p) for p in created)
                        def build_done(f, n=len(created), d=folder, fl=file_list):
                            ctk.CTkLabel(
                                f, text=f"Created {n} file(s) in:\n{d}",
                                font=make_font(FS_SM), text_color=TEXT_PRIMARY,
                                wraplength=320, justify="left"
                            ).pack(anchor="w", pady=(0, 8))
                            ctk.CTkLabel(
                                f, text=fl, font=make_font(FS_XS),
                                text_color=TEXT_MUTED, justify="left"
                            ).pack(anchor="w")
                        self._show_action_panel("Export Complete", build_done)
                    else:
                        self._hide_action_panel()
                except Exception as e:
                    def build_err(f, err=str(e)):
                        ctk.CTkLabel(
                            f, text=f"Failed to export:\n{err}",
                            font=make_font(FS_SM), text_color=RED_TEXT,
                            wraplength=320, justify="left"
                        ).pack(anchor="w")
                    self._show_action_panel("Export Error", build_err)
                    import traceback
                    traceback.print_exc()

            ctk.CTkButton(
                content, text="Export",
                command=do_export,
                fg_color=ACCENT, hover_color=ACCENT_HOVER,
                text_color=TEXT_ON_ACCENT,
                font=make_font(FS_SM, "bold"),
                height=34, corner_radius=RADIUS_SM
            ).pack(fill="x")

        self._show_action_panel("Export Calendar Files", build)

    # ── Email Schedule Dialog ────────────────────────────────────────────────
    def email_schedule_dialog(self):
        """
        Attempt to email each agent their .ics via Outlook COM.
        First-time use asks the user to verify Outlook works.
        Falls back to folder export if COM is unavailable.
        """
        if not self.generator:
            return

        settings = self.config_manager.get_settings()

        # ── Gate 1: Platform check ───────────────────────────────────────
        if not is_outlook_available():
            def build(f):
                ctk.CTkLabel(
                    f, text="Outlook COM automation requires Windows with "
                    "pywin32 installed.",
                    font=make_font(FS_SM), text_color=TEXT_PRIMARY,
                    wraplength=320, justify="left"
                ).pack(anchor="w", pady=(0, 10))
                ctk.CTkLabel(
                    f, text="You can still export .ics files and distribute "
                    "them manually using the Export Calendar button.",
                    font=make_font(FS_SM), text_color=TEXT_MUTED,
                    wraplength=320, justify="left"
                ).pack(anchor="w")
            self._show_action_panel("Outlook Not Available", build)
            return

        # ── Gate 2: First-time verification ──────────────────────────────
        if not settings.get("email_method_verified"):
            self._show_email_verification_panel(settings)
            return

        # ── Gate 3: Build and send ───────────────────────────────────────
        self._do_email_send()

    def _show_email_verification_panel(self, settings):
        """First-time inline panel: verify Outlook COM works for this user."""
        def build(f):
            ctk.CTkLabel(
                f, text="This feature uses the desktop Outlook app to "
                "send schedules.\n\nDo you use Outlook on this PC, and "
                "does it work correctly?\n\n(If you aren't sure, we can "
                "run a quick test.)",
                font=make_font(FS_SM), text_color=TEXT_PRIMARY,
                wraplength=320, justify="left"
            ).pack(anchor="w", pady=(0, 14))

            def on_yes():
                settings["email_method"] = "outlook_com"
                settings["email_method_verified"] = True
                self.config_manager.save_settings(settings)
                self._hide_action_panel()
                self._do_email_send()

            def on_no():
                self._hide_action_panel()
                def build_disabled(f2):
                    ctk.CTkLabel(
                        f2, text="No problem! You can still export .ics files "
                        "with the Export Calendar button and share them manually.",
                        font=make_font(FS_SM), text_color=TEXT_PRIMARY,
                        wraplength=320, justify="left"
                    ).pack(anchor="w")
                self._show_action_panel("Email Disabled", build_disabled)

            def on_idk():
                self._hide_action_panel()
                try:
                    send_via_outlook(
                        recipients=[],
                        subject="[Test] Schedule Builder — Outlook Check",
                        body_html="<p>If you can see this, Outlook COM is working!</p>"
                    )
                    self._show_email_test_result_panel(settings)
                except Exception as e:
                    def build_err(f2):
                        ctk.CTkLabel(
                            f2, text=f"Could not open Outlook:\n{str(e)}",
                            font=make_font(FS_SM), text_color=RED_TEXT,
                            wraplength=320, justify="left"
                        ).pack(anchor="w", pady=(0, 8))
                        ctk.CTkLabel(
                            f2, text="Use Export Calendar to export .ics files manually.",
                            font=make_font(FS_SM), text_color=TEXT_MUTED,
                            wraplength=320, justify="left"
                        ).pack(anchor="w")
                    self._show_action_panel("Outlook Error", build_err)

            ctk.CTkButton(
                f, text="Yes, I use Outlook", command=on_yes,
                fg_color=ACCENT, hover_color=ACCENT_HOVER,
                text_color=TEXT_ON_ACCENT,
                font=make_font(FS_SM, "bold"),
                height=34, corner_radius=RADIUS_SM
            ).pack(fill="x", pady=(0, 6))

            ctk.CTkButton(
                f, text="No", command=on_no,
                fg_color=SURFACE_ALT, hover_color=BORDER,
                text_color=TEXT_PRIMARY,
                border_width=1, border_color=BORDER,
                font=make_font(FS_SM, "bold"),
                height=34, corner_radius=RADIUS_SM
            ).pack(fill="x", pady=(0, 6))

            ctk.CTkButton(
                f, text="I Don't Know — Run Test", command=on_idk,
                fg_color=SURFACE_ALT, hover_color=BORDER,
                text_color=TEXT_PRIMARY,
                border_width=1, border_color=BORDER,
                font=make_font(FS_SM, "bold"),
                height=34, corner_radius=RADIUS_SM
            ).pack(fill="x")

        self._show_action_panel("Email Setup", build)

    def _show_email_test_result_panel(self, settings):
        """After the Outlook test draft, ask if it worked."""
        def build(f):
            ctk.CTkLabel(
                f, text="A test email draft should have opened in Outlook.\n\n"
                "Did it open successfully?",
                font=make_font(FS_SM), text_color=TEXT_PRIMARY,
                wraplength=320, justify="left"
            ).pack(anchor="w", pady=(0, 14))

            def on_worked():
                settings["email_method"] = "outlook_com"
                settings["email_method_verified"] = True
                self.config_manager.save_settings(settings)
                self._hide_action_panel()
                self._do_email_send()

            def on_didnt():
                self._hide_action_panel()
                def build_alt(f2):
                    ctk.CTkLabel(
                        f2, text="That's okay! Use Export Calendar to save "
                        ".ics files and share them manually.",
                        font=make_font(FS_SM), text_color=TEXT_PRIMARY,
                        wraplength=320, justify="left"
                    ).pack(anchor="w")
                self._show_action_panel("Email Disabled", build_alt)

            ctk.CTkButton(
                f, text="Yes, it worked!", command=on_worked,
                fg_color=ACCENT, hover_color=ACCENT_HOVER,
                text_color=TEXT_ON_ACCENT,
                font=make_font(FS_SM, "bold"),
                height=34, corner_radius=RADIUS_SM
            ).pack(fill="x", pady=(0, 6))

            ctk.CTkButton(
                f, text="No, it didn't work", command=on_didnt,
                fg_color=SURFACE_ALT, hover_color=BORDER,
                text_color=TEXT_PRIMARY,
                border_width=1, border_color=BORDER,
                font=make_font(FS_SM, "bold"),
                height=34, corner_radius=RADIUS_SM
            ).pack(fill="x")

        self._show_action_panel("Did It Work?", build)

    def _do_email_send(self):
        """
        Generate per-agent .ics files into a temp folder, then open
        an Outlook draft addressed to everyone with attachments.
        """
        import tempfile

        month_name = self.month_var.get()
        year = int(self.year_var.get())
        month_num = list(calendar.month_name).index(month_name)

        # Check if any employees have emails
        recipients = [(e["name"], e.get("email", "")) for e in self.employees]
        has_emails = any(email for _, email in recipients)

        if not has_emails:
            def build(f):
                ctk.CTkLabel(
                    f, text="None of your team members have email addresses "
                    "configured.",
                    font=make_font(FS_SM), text_color=TEXT_PRIMARY,
                    wraplength=320, justify="left"
                ).pack(anchor="w", pady=(0, 8))
                ctk.CTkLabel(
                    f, text="Add emails in Manage Team, or use Export Calendar "
                    "to save .ics files and distribute them manually.",
                    font=make_font(FS_SM), text_color=TEXT_MUTED,
                    wraplength=320, justify="left"
                ).pack(anchor="w")
            self._show_action_panel("No Email Addresses", build)
            return

        try:
            # Create temp folder for the .ics files and Excel
            tmp_dir = tempfile.mkdtemp(prefix="schedule_ics_")
            created = generate_all_ics(
                self.generator, self.employees, self.edited_assignments,
                year, month_num, tmp_dir, mode="separate"
            )

            if not created:
                def build(f):
                    ctk.CTkLabel(
                        f, text="No calendar events to send.",
                        font=make_font(FS_SM), text_color=TEXT_MUTED,
                        wraplength=320, justify="left"
                    ).pack(anchor="w")
                self._show_action_panel("No Events", build)
                return

            # Generate Excel file and add to attachments
            excel_path = os.path.join(tmp_dir, f"Schedule_{month_name}_{year}.xlsx")
            self._create_excel_file(excel_path)
            created.append(excel_path)

            subject = f"Helpdesk Schedule — {month_name} {year}"
            body_html = build_email_body(month_name, year, self.employees)

            send_via_outlook(
                recipients=recipients,
                subject=subject,
                body_html=body_html,
                attachment_paths=created
            )

            def build(f):
                num_ics = len(created) - 1  # Subtract 1 for the Excel file
                ctk.CTkLabel(
                    f, text=f"An Outlook email draft has been opened with:\n"
                    f"• {num_ics} calendar file(s) (.ics)\n"
                    f"• 1 Excel schedule (.xlsx)\n\n"
                    f"Review and hit Send when ready!",
                    font=make_font(FS_SM), text_color=TEXT_PRIMARY,
                    wraplength=320, justify="left"
                ).pack(anchor="w")
            self._show_action_panel("Draft Opened", build)

        except RuntimeError as e:
            def build(f, err=str(e)):
                ctk.CTkLabel(
                    f, text=err, font=make_font(FS_SM),
                    text_color=RED_TEXT, wraplength=320, justify="left"
                ).pack(anchor="w")
            self._show_action_panel("Outlook Error", build)
        except Exception as e:
            def build(f, err=str(e)):
                ctk.CTkLabel(
                    f, text=f"Something went wrong:\n{err}",
                    font=make_font(FS_SM), text_color=RED_TEXT,
                    wraplength=320, justify="left"
                ).pack(anchor="w", pady=(0, 8))
                ctk.CTkLabel(
                    f, text="Try using Export Calendar instead.",
                    font=make_font(FS_SM), text_color=TEXT_MUTED,
                    wraplength=320, justify="left"
                ).pack(anchor="w")
            self._show_action_panel("Email Error", build)
            import traceback
            traceback.print_exc()

    def _reset_email_settings(self):
        """Reset the Outlook COM verification so the user is asked again."""
        settings = self.config_manager.get_settings()
        settings["email_method"] = None
        settings["email_method_verified"] = False
        self.config_manager.save_settings(settings)
        def build(f):
            ctk.CTkLabel(
                f, text="Email settings have been cleared.\n\n"
                "You'll be asked to verify Outlook next time you use "
                "Email Schedule.",
                font=make_font(FS_SM), text_color=TEXT_PRIMARY,
                wraplength=320, justify="left"
            ).pack(anchor="w")
        self._show_action_panel("Settings Reset", build)

    # ── Excel Export ──────────────────────────────────────────────────────────
    def _create_excel_file(self, filepath):
        """Create an Excel file at the given path with the schedule data."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = f"{self.month_var.get()} Schedule"[:31]

        # Styles
        fill_header = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        fill_phone  = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        fill_fd     = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        fill_empty  = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        fill_error  = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        font_header = Font(color="FFFFFF", bold=True)
        font_label  = Font(bold=True)
        align_c     = Alignment(horizontal="center", vertical="center")

        weeks = self.generator.get_weekly_matrix()
        current_row = 1

        for week_idx, week_data in enumerate(weeks):
            ws.cell(row=current_row, column=1,
                    value=f"Week {week_idx + 1}").font = font_label
            days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

            for col_i, (day_name, date_str) in enumerate(zip(days, week_data["dates"])):
                cell = ws.cell(row=current_row, column=col_i + 2)
                cell.value = f"{day_name}\n{date_str}" if date_str else day_name
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

            current_row += 1

            for s_idx in ALL_SHIFTS:
                label_cell = ws.cell(row=current_row, column=1,
                                     value=SHIFT_NAMES[s_idx])
                label_cell.font = font_label
                label_cell.alignment = Alignment(vertical="center")

                for col_i, worker in enumerate(week_data["matrix"][s_idx]):
                    final = worker
                    # Apply manual edits to Excel export
                    if (week_idx, col_i, s_idx) in self.edited_assignments:
                        final = self.edited_assignments[(week_idx, col_i, s_idx)]
                    cell = ws.cell(row=current_row, column=col_i + 2)
                    cell.value = final
                    cell.alignment = align_c

                    if not week_data["dates"][col_i]:
                        cell.fill = fill_empty
                    elif not final:
                        cell.fill = fill_error
                    elif s_idx in PHONE_SHIFTS:
                        cell.fill = fill_phone
                    else:
                        cell.fill = fill_fd

                current_row += 1

            current_row += 2

        # Fairness audit
        current_row += 1
        ws.cell(row=current_row, column=1,
                value="Fairness Audit").font = Font(size=14, bold=True)
        current_row += 1

        for col_i, header in enumerate(["Employee", "Phone Shifts", "Front Desk", "Total"]):
            cell = ws.cell(row=current_row, column=col_i + 1, value=header)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = align_c
        current_row += 1

        for stat in self._get_stats_with_edits(self.generator):
            ws.cell(row=current_row, column=1, value=stat["name"]).alignment = align_c
            ws.cell(row=current_row, column=2, value=stat["phone"]).alignment = align_c
            ws.cell(row=current_row, column=3, value=stat["fd"]).alignment = align_c
            ws.cell(row=current_row, column=4, value=stat["total"]).alignment = align_c
            current_row += 1

        ws.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15

        wb.save(filepath)

    def save_excel(self):
        if not self.generator:
            return
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"Schedule_{self.month_var.get()}_{self.year_var.get()}.xlsx"
            )
            if not filename:
                return

            self._create_excel_file(filename)
            messagebox.showinfo("Saved", f"Schedule saved to:\n{filename}")

            import subprocess
            abs_path = os.path.abspath(filename)
            if os.path.isfile(abs_path):
                if sys.platform == "darwin":
                    subprocess.run(["open", abs_path], check=False)
                elif sys.platform == "win32":
                    os.startfile(abs_path)
                else:
                    subprocess.run(["xdg-open", abs_path], check=False)

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export: {str(e)}")
            import traceback
            traceback.print_exc()


if __name__ == "__main__":
    app = ScheduleAppV2()
    app.mainloop()
